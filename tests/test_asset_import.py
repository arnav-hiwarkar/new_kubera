"""Bulk import of pre-existing assets from Excel/CSV."""
import io

import openpyxl
import pytest
from httpx import AsyncClient

from tests.asset_helpers import admin_headers

ASSETS = "/api/v1/assets"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER = ["Asset name", "Category", "Subcategory", "Original cost",
          "Purchase date", "Put-to-use date", "Capitalization date",
          "Opening accumulated depreciation", "Opening WDV (books)", "Opening WDV (tax)",
          "Useful life months", "Dep method", "Residual %", "IT block code",
          "Branch", "Location", "Department", "Cost centre", "Custodian name",
          "Serial number", "Remarks"]


def xlsx_of(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


GOOD_ROW = ["Tata Ace", "Motor vehicles", "Motor cars (other than those used in a hire business)",
            850000, "2022-06-10", "2022-06-20", "2022-06-30",
            200000, 650000, 610000, None, None, None, None,
            None, None, None, None, "R Kumar", "DL1AB1234", "Bought pre-register"]


def upload(content, filename="assets.xlsx"):
    return {"file": (filename, content, XLSX_MIME)}


@pytest.mark.asyncio
async def test_template_downloads_with_expected_columns(client: AsyncClient):
    AH = await admin_headers(client, "imp_tpl@a.com")
    resp = await client.get(f"{ASSETS}/import/template", headers=AH)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MIME)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Instructions", "Assets"]
    assert [c.value for c in wb["Assets"][1]] == HEADER


@pytest.mark.asyncio
async def test_import_creates_drafts_and_is_atomic(client: AsyncClient):
    AH = await admin_headers(client, "imp_ok@a.com")

    resp = await client.post(f"{ASSETS}/import", files=upload(xlsx_of([GOOD_ROW, GOOD_ROW])),
                             headers=AH)
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["created_count"] == 2 and body["first_asset_id"]

    listing = (await client.get(f"{ASSETS}", headers=AH)).json()
    assert len(listing) == 2
    assert all(a["is_pre_cutover"] and a["lifecycle_status"] == "draft" for a in listing)

    # One bad row aborts everything.
    bad = list(GOOD_ROW)
    bad[0] = ""
    resp2 = await client.post(f"{ASSETS}/import", files=upload(xlsx_of([GOOD_ROW, bad])), headers=AH)
    assert resp2.status_code == 422
    errs = resp2.json()["detail"]
    assert isinstance(errs, list) and errs[0]["row"] == 3
    assert len((await client.get(f"{ASSETS}", headers=AH)).json()) == 2  # unchanged


@pytest.mark.asyncio
async def test_csv_with_case_insensitive_categories(client: AsyncClient):
    AH = await admin_headers(client, "imp_csv@a.com")
    csv_content = (
        "Asset name,Category,Subcategory,Original cost,,,,,,,,,,,,,,,,,,,\n"
        "old chair,furniture and fittings,general furniture and fittings,3000,,,,,,,,,,,,,,,,,,,\n"
    ).encode("utf-8")
    resp = await client.post(
        f"{ASSETS}/import", files=upload(csv_content, "assets.csv"), headers=AH)
    assert resp.status_code == 201, resp.text
    listing = (await client.get(f"{ASSETS}", headers=AH)).json()
    assert listing[0]["useful_life_months"] == 120  # category default applied
