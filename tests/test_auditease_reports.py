"""Integration tests for AuditEase statutory reports and export endpoints."""
import io
import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy import select

from app.models.auditease import AuditEngagement
from app.models.company import Company, CompanyUser
from app.services import trial_balance_query as tbq
from app.services.reporting.auditease_reports import (
    build_balance_sheet,
    build_profit_and_loss,
    build_notes_to_accounts,
    build_extended_trial_balance,
    build_adjusting_entries,
)
from tests.conftest import create_test_company, get_company_token, TestSessionLocal
from tests.test_auditease import import_tb, make_engagement


async def setup_mapped_engagement(client: AsyncClient, email: str = "admin_rep@testco.com"):
    company_data = await create_test_company(client, name="AuditEase Reporting Test Co", email=email)
    token = await get_company_token(client, email=email)
    headers = {"Authorization": f"Bearer {token}"}

    eng_id = await make_engagement(client, headers, label="FY 2024-25")

    csv_content = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"1001,Equity Share Capital,0,0,100000,-100000\n"
        b"2001,HDFC Term Loan,0,0,50000,-50000\n"
        b"3001,Plant and Machinery,120000,0,0,120000\n"
        b"4001,SBI Current Account,0,50000,20000,30000\n"
        b"5001,Sales Revenue,0,0,80000,-80000\n"
        b"6001,Salaries Expense,0,80000,0,80000\n"
    )
    res = await import_tb(client, eng_id, headers, csv=csv_content)
    assert res.status_code == 200, res.text

    grp_res = await client.get("/api/v1/auditease/ledger-groups", headers=headers)
    assert grp_res.status_code == 200, grp_res.text
    groups = grp_res.json()
    group_by_name = {g["name"]: g for g in groups}

    tb_res = await client.get(f"/api/v1/auditease/engagements/{eng_id}/trial-balance", headers=headers)
    assert tb_res.status_code == 200, tb_res.text
    accounts = tb_res.json()["accounts"]
    acc_by_name = {a["ledger_name"]: a for a in accounts}

    mappings = [
        ("Equity Share Capital", "Share Capital"),
        ("HDFC Term Loan", "Long-term Borrowings"),
        ("Plant and Machinery", "Property, Plant and Equipment"),
        ("SBI Current Account", "Cash and Cash Equivalents"),
        ("Sales Revenue", "Revenue from Operations"),
        ("Salaries Expense", "Employee Benefits Expense"),
    ]

    for ledger_name, group_name in mappings:
        if ledger_name in acc_by_name and group_name in group_by_name:
            map_res = await client.post(
                f"/api/v1/auditease/engagements/{eng_id}/ledgers/{acc_by_name[ledger_name]['id']}/map",
                json={"group_id": group_by_name[group_name]["id"]},
                headers=headers,
            )
            assert map_res.status_code == 200, map_res.text

    return {
        "headers": headers,
        "eng_id": eng_id,
        "email": email,
        "accounts": accounts,
        "acc_by_name": acc_by_name,
        "groups": groups,
        "group_by_name": group_by_name,
    }


@pytest.mark.asyncio
async def test_auditease_balance_sheet_invariants(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="bs_inv@testco.com")
    headers = ctx["headers"]
    eng_id = ctx["eng_id"]
    email = ctx["email"]

    # 1. Preview API response check
    preview_res = await client.get(f"/api/v1/auditease/engagements/{eng_id}/reports/preview", headers=headers)
    assert preview_res.status_code == 200, preview_res.text
    data = preview_res.json()
    totals = data["totals"]
    balance_check = data["balance_check"]
    assert totals["assets"] == balance_check["assets"]
    assert balance_check["assets"] == balance_check["liabilities_plus_equity"]
    assert balance_check["balanced"] is True
    assert totals["assets"] > 0
    assert totals["assets"] == 150000.0

    # 2. Document builders structure check
    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        company = await session.get(Company, user.company_id)
        eng = await session.get(AuditEngagement, eng_id)
        figures = await tbq.load_engagement_figures(session, user.company_id, eng_id)
        warnings = tbq.view_warnings(figures.figures, figures.summary, eng.tb_sign_convention)

    doc = build_balance_sheet(figures.figures, figures.summary, company.name, eng.period_label, "absolute", warnings)

    sec_eq = doc.sections[0]
    sec_assets = doc.sections[1]

    # Every sub-section total equals the sum of its rows
    for sec in (*sec_eq.children, *sec_assets.children):
        if sec.total and sec.rows:
            expected_sum = sum(r.cells["amount"] for r in sec.rows)
            assert sec.total.cells["amount"] == expected_sum

    # Section totals equal sum of child sections
    eq_children_sum = sum(c.total.cells["amount"] for c in sec_eq.children if c.total)
    assert sec_eq.total.cells["amount"] == eq_children_sum

    assets_children_sum = sum(c.total.cells["amount"] for c in sec_assets.children if c.total)
    assert sec_assets.total.cells["amount"] == assets_children_sum

    # Balance Sheet balances
    assert sec_assets.total.cells["amount"] == sec_eq.total.cells["amount"]
    assert sec_assets.total.cells["amount"] > 0


@pytest.mark.asyncio
async def test_auditease_profit_and_loss_invariants(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="pl_inv@testco.com")
    email = ctx["email"]
    eng_id = ctx["eng_id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        company = await session.get(Company, user.company_id)
        eng = await session.get(AuditEngagement, eng_id)
        figures = await tbq.load_engagement_figures(session, user.company_id, eng_id)
        warnings = tbq.view_warnings(figures.figures, figures.summary, eng.tb_sign_convention)

    doc = build_profit_and_loss(figures.figures, figures.summary, company.name, eng.period_label, "absolute", warnings)
    sec_inc, sec_exp, sec_profit = doc.sections

    tot_inc = sum(r.cells["amount"] for r in sec_inc.rows)
    assert sec_inc.total.cells["amount"] == tot_inc

    tot_exp = sum(r.cells["amount"] for r in sec_exp.rows)
    assert sec_exp.total.cells["amount"] == tot_exp

    assert sec_profit.total.cells["amount"] == tot_inc - tot_exp
    assert sec_profit.total.cells["amount"] == figures.summary.net_profit


@pytest.mark.asyncio
async def test_auditease_notes_to_accounts_invariants(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="notes_inv@testco.com")
    email = ctx["email"]
    eng_id = ctx["eng_id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        company = await session.get(Company, user.company_id)
        eng = await session.get(AuditEngagement, eng_id)
        figures = await tbq.load_engagement_figures(session, user.company_id, eng_id)
        warnings = tbq.view_warnings(figures.figures, figures.summary, eng.tb_sign_convention)

    doc = build_notes_to_accounts(figures.figures, figures.summary, company.name, eng.period_label, "absolute", warnings)

    for sec in doc.sections:
        if sec.total and sec.rows:
            assert sec.total.cells["final"] == sum(r.cells["final"] for r in sec.rows)
            assert sec.total.cells["closing"] == sum(r.cells["closing"] for r in sec.rows)


@pytest.mark.asyncio
async def test_auditease_extended_trial_balance_invariants(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="etb_inv@testco.com")
    email = ctx["email"]
    eng_id = ctx["eng_id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        company = await session.get(Company, user.company_id)
        eng = await session.get(AuditEngagement, eng_id)
        figures = await tbq.load_engagement_figures(session, user.company_id, eng_id)
        warnings = tbq.view_warnings(figures.figures, figures.summary, eng.tb_sign_convention)

    doc = build_extended_trial_balance(figures.figures, figures.summary, company.name, eng.period_label, "absolute", warnings)
    total_cells = doc.sections[0].total.cells

    assert total_cells["unadj_dr"] == total_cells["unadj_cr"]
    assert total_cells["adj_dr"] == total_cells["adj_cr"]
    assert total_cells["adj_tb_dr"] == total_cells["adj_tb_cr"]


@pytest.mark.asyncio
async def test_auditease_adjusting_entries_ledger_name(client: AsyncClient):
    """Adjusting entries report must display ledger name, not UUID string.
    
    Expected to FAIL: defect R6a (EntryLine has no ledger_name attribute).
    """
    ctx = await setup_mapped_engagement(client, email="adj_name@testco.com")
    headers = ctx["headers"]
    eng_id = ctx["eng_id"]
    email = ctx["email"]
    acc_by_name = ctx["acc_by_name"]

    ppe_id = acc_by_name["Plant and Machinery"]["id"]
    exp_id = acc_by_name["Salaries Expense"]["id"]

    # Register and invite auditor to propose adjusting entry
    aud_email = "aud_adj@testco.com"
    reg_res = await client.post("/api/v1/auth/auditor/register", json={"email": aud_email, "password": "pass1234", "name": "Auditor Adj"})
    assert reg_res.status_code == 201, reg_res.text
    resp_login = await client.post("/api/v1/auth/auditor/login", json={"email": aud_email, "password": "pass1234"})
    assert resp_login.status_code == 200, resp_login.text
    aud_headers = {"Authorization": f"Bearer {resp_login.json()['access_token']}"}

    inv_res = await client.post(f"/api/v1/auditease/engagements/{eng_id}/invite-auditor", json={"email": aud_email}, headers=headers)
    assert inv_res.status_code == 200, inv_res.text

    acc_res = await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)
    assert acc_res.status_code == 200, acc_res.text

    # Post adjusting entry from auditor
    adj_res = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/entries",
        json={
            "description": "Reclassify expense to capital",
            "lines": [
                {"ledger_id": ppe_id, "side": "debit", "amount": 5000.0},
                {"ledger_id": exp_id, "side": "credit", "amount": 5000.0},
            ],
        },
        headers=aud_headers,
    )
    assert adj_res.status_code == 201, adj_res.text
    entry_id = adj_res.json()["id"]

    # Approve adjusting entry as company
    appr_res = await client.patch(
        f"/api/v1/auditease/entries/{entry_id}/approve",
        json={"status": "approved"},
        headers=headers,
    )
    assert appr_res.status_code == 200, appr_res.text

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        company = await session.get(Company, user.company_id)
        eng = await session.get(AuditEngagement, eng_id)
        figures = await tbq.load_engagement_figures(session, user.company_id, eng_id)
        warnings = tbq.view_warnings(figures.figures, figures.summary, eng.tb_sign_convention)

    doc = build_adjusting_entries(figures.approved_entries, company.name, eng.period_label, "absolute", warnings)

    ledger_names_rendered = [row.cells["ledger"] for row in doc.sections[0].rows]
    assert "Plant and Machinery" in ledger_names_rendered, f"Expected ledger name in report rows, got: {ledger_names_rendered}"
    assert "Salaries Expense" in ledger_names_rendered, f"Expected ledger name in report rows, got: {ledger_names_rendered}"


@pytest.mark.asyncio
async def test_auditease_balance_sheet_reconciliation_top_group(client: AsyncClient):
    """Balance Sheet printed rows must sum to TOTAL ASSETS when a custom subgroup is mapped.
    
    Expected to FAIL: defect R7a (_get_group_node_amount only checks hardcoded Schedule III groups).
    """
    email = "bs_recon@testco.com"
    await create_test_company(client, name="Reconciliation Co", email=email)
    token = await get_company_token(client, email=email)
    headers = {"Authorization": f"Bearer {token}"}
    eng_id = await make_engagement(client, headers, label="FY 2024-25")

    csv_content = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"1001,Equity Share Capital,0,0,50000,-50000\n"
        b"3001,Custom Direct Asset,50000,0,0,50000\n"
    )
    imp_res = await import_tb(client, eng_id, headers, csv=csv_content)
    assert imp_res.status_code == 200, imp_res.text
    accs = imp_res.json()["accounts"]
    share_acc = next(a for a in accs if a["ledger_name"] == "Equity Share Capital")
    asset_acc = next(a for a in accs if a["ledger_name"] == "Custom Direct Asset")

    grp_res = await client.get("/api/v1/auditease/ledger-groups", headers=headers)
    assert grp_res.status_code == 200, grp_res.text
    groups = grp_res.json()
    share_cap_group = next(g for g in groups if g["name"] == "Share Capital")
    assets_top = next(g for g in groups if g["name"] == "Assets" and g["parent_id"] is None)

    # Map share capital to Share Capital group
    res1 = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/ledgers/{share_acc['id']}/map",
        json={"group_id": share_cap_group["id"]},
        headers=headers,
    )
    assert res1.status_code == 200, res1.text

    # Create a custom subgroup under Assets and map asset to it
    res_grp = await client.post(
        "/api/v1/auditease/ledger-groups",
        json={"name": "Special Equipment", "parent_id": assets_top["id"]},
        headers=headers,
    )
    assert res_grp.status_code == 201, res_grp.text
    custom_leaf = res_grp.json()

    res_map = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/ledgers/{asset_acc['id']}/map",
        json={"group_id": custom_leaf["id"]},
        headers=headers,
    )
    assert res_map.status_code == 200, res_map.text

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        figures = await tbq.load_engagement_figures(session, user.company_id, eng_id)

    doc = build_balance_sheet(figures.figures, figures.summary, "Reconciliation Co", "FY 2024-25", "absolute", ())

    # Sum the printed rows in the Assets section (doc.sections[1])
    printed_rows_sum = sum(
        sum(r.cells["amount"] for r in sub_sec.rows)
        for sub_sec in doc.sections[1].children
    )
    total_assets_printed = doc.sections[1].total.cells["amount"]

    assert printed_rows_sum == total_assets_printed, (
        f"Printed asset rows ({printed_rows_sum}) do not sum to TOTAL ASSETS ({total_assets_printed})"
    )


@pytest.mark.asyncio
async def test_auditease_reports_exports_xlsx_pdf_html(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="exports_all@testco.com")
    headers = ctx["headers"]
    eng_id = ctx["eng_id"]

    report_keys = [
        "balance_sheet",
        "profit_and_loss",
        "notes_to_accounts",
        "trial_balance_detailed",
        "trial_balance_summary",
        "extended_trial_balance",
        "adjusting_entries",
        "ledger_mapping",
        "exceptions",
    ]

    for key in report_keys:
        # Export XLSX
        xlsx_res = await client.get(
            f"/api/v1/auditease/engagements/{eng_id}/reports/{key}/export?format=xlsx&units=absolute",
            headers=headers,
        )
        assert xlsx_res.status_code == 200, f"Failed xlsx export for {key}: {xlsx_res.text}"
        assert "spreadsheetml" in xlsx_res.headers["content-type"]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_res.content))
        assert len(wb.sheetnames) >= 1

        # Export PDF
        pdf_res = await client.get(
            f"/api/v1/auditease/engagements/{eng_id}/reports/{key}/export?format=pdf&units=thousands",
            headers=headers,
        )
        assert pdf_res.status_code == 200, f"Failed pdf export for {key}: {pdf_res.text}"
        assert pdf_res.headers["content-type"] == "application/pdf"
        assert pdf_res.content.startswith(b"%PDF")

        # Preview HTML
        html_res = await client.get(
            f"/api/v1/auditease/engagements/{eng_id}/reports/{key}/preview-html?units=lakhs",
            headers=headers,
        )
        assert html_res.status_code == 200, f"Failed preview-html for {key}: {html_res.text}"
        data = html_res.json()
        assert "html" in data
        assert "AuditEase Reporting Test Co" in data["html"]


@pytest.mark.asyncio
async def test_auditease_report_pack(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="pack_test@testco.com")
    headers = ctx["headers"]
    eng_id = ctx["eng_id"]

    pack_xlsx = await client.get(
        f"/api/v1/auditease/engagements/{eng_id}/reports/pack?format=xlsx&units=absolute",
        headers=headers,
    )
    assert pack_xlsx.status_code == 200, pack_xlsx.text
    pack_wb = openpyxl.load_workbook(io.BytesIO(pack_xlsx.content))
    assert "Cover" in pack_wb.sheetnames
    assert "Balance Sheet" in pack_wb.sheetnames

    pack_pdf = await client.get(
        f"/api/v1/auditease/engagements/{eng_id}/reports/pack?format=pdf&units=absolute",
        headers=headers,
    )
    assert pack_pdf.status_code == 200, pack_pdf.text
    assert pack_pdf.content.startswith(b"%PDF")
    assert len(pack_pdf.content) > 3000


@pytest.mark.asyncio
async def test_auditease_report_archive(client: AsyncClient):
    ctx = await setup_mapped_engagement(client, email="arch_test@testco.com")
    headers = ctx["headers"]
    eng_id = ctx["eng_id"]

    archive_res = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/reports/archive?report_key=pack&format=pdf&units=absolute",
        headers=headers,
    )
    assert archive_res.status_code == 200, archive_res.text
    res_data = archive_res.json()
    assert "id" in res_data
    assert "/api/v1/docvault/documents/" in res_data["url"]
