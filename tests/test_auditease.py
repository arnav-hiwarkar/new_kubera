import json
import uuid

import pytest
from httpx import AsyncClient

from tests.conftest import create_test_company, get_company_token
from app.models.auditease import EngagementStatus, GrantStatus, AuditEntryStatus, RequestStatus, QueryStatus

# --- Trial-balance import fixtures/helpers -------------------------------------

TB_CSV = (
    b"Code,Name,Opening,Debit,Credit,Closing\n"
    b"A1,Cash,100,50,0,150\n"
    b"L1,Loan,-100,0,50,-150\n"
)
TB_MAP = {
    "ledger_code": "Code",
    "ledger_name": "Name",
    "opening_balance": "Opening",
    "debit": "Debit",
    "credit": "Credit",
    "closing_balance": "Closing",
}


async def make_engagement(client, headers, label="FY24"):
    resp = await client.post("/api/v1/auditease/engagements", json={"period_label": label}, headers=headers)
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def import_tb(client, eng_id, headers, csv=TB_CSV, cmap=TB_MAP, sheet=None,
                    convention=None, header_row=None, confirm=None):
    data = {"column_map": json.dumps(cmap)}
    if sheet is not None:
        data["sheet"] = sheet
    if convention is not None:
        data["sign_convention"] = convention
    if header_row is not None:
        data["header_row"] = str(header_row)
    if confirm is not None:
        data["confirm"] = "true" if confirm else "false"
    return await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/trial-balance/import",
        data=data,
        files={"file": ("tb.csv", csv, "text/csv")},
        headers=headers,
    )


async def preview_tb(client, eng_id, headers, csv=TB_CSV, cmap=TB_MAP,
                     convention=None, header_row=None):
    data = {"column_map": json.dumps(cmap)}
    if convention is not None:
        data["sign_convention"] = convention
    if header_row is not None:
        data["header_row"] = str(header_row)
    return await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/trial-balance/preview",
        data=data,
        files={"file": ("tb.csv", csv, "text/csv")},
        headers=headers,
    )


async def get_tb(client, eng_id, headers):
    """GET /trial-balance now returns {accounts, totals, ...} rather than a bare array."""
    resp = await client.get(
        f"/api/v1/auditease/engagements/{eng_id}/trial-balance", headers=headers
    )
    return resp


# --- Tests ---------------------------------------------------------------------

@pytest.mark.asyncio
async def test_trial_balance_import_flow(client: AsyncClient):
    await create_test_company(client, email="tb@a.com", password="pass1234")
    token = await get_company_token(client, email="tb@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {token}"}

    eng_id = await make_engagement(client, headers)

    # Step 1: inspect returns sheet headers + preview
    resp = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/trial-balance/inspect",
        files={"file": ("tb.csv", TB_CSV, "text/csv")},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    sheets = resp.json()["sheets"]
    assert sheets[0]["headers"] == ["Code", "Name", "Opening", "Debit", "Credit", "Closing"]
    assert len(sheets[0]["preview_rows"]) == 2

    # Step 2: import with column map
    resp = await import_tb(client, eng_id, headers)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["imported"] == 2
    assert body["skipped"] == 0
    # `balanced` now means "the trial balance sums to zero", the same definition the
    # grid and the report use -- not "total debit column == total credit column".
    assert body["balanced"] is True
    # Closings are 150 / -150, so the signed convention is provable from the file.
    assert body["sign_convention"] == "signed"
    assert body["diagnostics"]["convention_confidence"] == "proven"
    assert body["diagnostics"]["closing_sums_to_zero"] is True
    assert [a["closing_net_debit"] for a in body["accounts"]] == [150.0, -150.0]

    # View the per-engagement TB
    resp = await get_tb(client, eng_id, headers)
    assert resp.status_code == 200
    view = resp.json()
    accounts = view["accounts"]
    assert len(accounts) == 2
    assert all(a["engagement_id"] == eng_id for a in accounts)
    assert view["totals"]["balanced"] is True
    assert view["totals"]["difference"] == 0.0
    assert view["sign_convention"] == "signed"


@pytest.mark.asyncio
async def test_tb_import_skips_bad_rows(client: AsyncClient):
    await create_test_company(client, email="bad@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='bad@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)

    csv = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"A1,Cash,100,50,0,150\n"
        b"B2,Bad,100,notanumber,0,150\n"   # non-numeric debit -> still a row error
        b"C3,Blanks,100,50,,150\n"          # blank credit -> now imports as zero
        b",,,,,\n"                          # fully blank -> dropped, not an error
    )
    resp = await import_tb(client, eng_id, headers, csv=csv)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    # Two good rows: the blank-credit row is no longer discarded.
    assert body["imported"] == 2
    assert body["diagnostics"]["rows_error"] == 1
    assert body["errors"][0]["row"] == 3
    assert body["diagnostics"]["rows_dropped_blank"] == 1
    blanks = next(a for a in body["accounts"] if a["ledger_name"] == "Blanks")
    assert blanks["credit"] == 0.0


@pytest.mark.asyncio
async def test_tb_import_flexible_layouts_and_preview_is_read_only(client: AsyncClient):
    await create_test_company(client, email="layouts@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='layouts@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)

    movements = b"Name,Debit,Credit\nCash,100,\nSales,,100\n"
    movement_map = {"ledger_name": "Name", "debit": "Debit", "credit": "Credit"}
    preview = await preview_tb(client, eng_id, headers, csv=movements, cmap=movement_map)
    assert preview.status_code == 200, preview.text
    assert preview.json()["diagnostics"]["detected_convention"] == "derived"
    assert (await get_tb(client, eng_id, headers)).json()["accounts"] == []

    imported = await import_tb(client, eng_id, headers, csv=movements, cmap=movement_map)
    assert imported.status_code == 200, imported.text
    assert sorted(a["closing_net_debit"] for a in imported.json()["accounts"]) == [-100.0, 100.0]

    eng2 = await make_engagement(client, headers, "FY25")
    single = b"Name,Closing\nCash,100\nSales,-100\n"
    single_map = {"ledger_name": "Name", "closing_balance": "Closing"}
    imported = await import_tb(client, eng2, headers, csv=single, cmap=single_map)
    assert imported.status_code == 200, imported.text
    assert imported.json()["sign_convention"] == "signed"


@pytest.mark.asyncio
async def test_tb_import_header_total_suffix_indian_and_oversized_diagnostics(client: AsyncClient):
    await create_test_company(client, email="formats@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='formats@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)
    csv = (
        b"Example Private Limited,,,\n"
        b"Trial Balance as at 31 March 2026,,,\n"
        b"Name,Debit,Credit,Closing\n"
        b"Cash,1,23,456 Dr,,123456 Dr\n"
    )
    # CSV commas must be quoted when they are part of the number.
    csv = csv.replace(b"Cash,1,23,456 Dr,,123456 Dr", b'Cash,"1,23,456 Dr",,123456 Dr')
    cmap = {"ledger_name": "Name", "debit": "Debit", "credit": "Credit", "closing_balance": "Closing"}
    inspected = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/trial-balance/inspect",
        files={"file": ("tb.csv", csv, "text/csv")}, headers=headers,
    )
    assert inspected.status_code == 200, inspected.text
    assert inspected.json()["sheets"][0]["header_row"] == 3

    too_large = b"Name,Closing\nHuge,10000000000000\n"
    preview = await preview_tb(
        client, eng_id, headers, csv=too_large,
        cmap={"ledger_name": "Name", "closing_balance": "Closing"},
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["diagnostics"]["rows_error"] == 1
    assert preview.json()["would_import"] == 0


@pytest.mark.asyncio
async def test_tb_reimport_replaces(client: AsyncClient):
    await create_test_company(client, email="re@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='re@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)

    await import_tb(client, eng_id, headers)
    await import_tb(client, eng_id, headers)  # second import updates in place, not appends
    resp = await get_tb(client, eng_id, headers)
    assert len(resp.json()["accounts"]) == 2


@pytest.mark.asyncio
async def test_tb_reimport_preserves_mappings_and_ids(client: AsyncClient):
    """Re-import upserts. The old code deleted every row and reinserted, throwing away
    all of the user's mapping work (and cascade-deleting audit entry lines)."""
    await create_test_company(client, email="upsert@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='upsert@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)

    resp = await import_tb(client, eng_id, headers)
    cash = next(a for a in resp.json()["accounts"] if a["ledger_name"] == "Cash")

    groups = (await client.get("/api/v1/auditease/ledger-groups", headers=headers)).json()
    leaf = next(g for g in groups if g["name"] == "Cash and Cash Equivalents")
    await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/ledgers/{cash['id']}/map",
        json={"group_id": leaf["id"]}, headers=headers,
    )

    # Re-import the same file with a changed figure.
    changed = TB_CSV.replace(b"A1,Cash,100,50,0,150", b"A1,Cash,100,75,0,175")
    resp = await import_tb(client, eng_id, headers, csv=changed)
    assert resp.status_code == 200, resp.text

    view = (await get_tb(client, eng_id, headers)).json()
    cash_after = next(a for a in view["accounts"] if a["ledger_name"] == "Cash")
    assert cash_after["id"] == cash["id"]                       # identity preserved
    assert cash_after["mapped_group_id"] == leaf["id"]          # mapping survived
    assert cash_after["closing_net_debit"] == 175.0             # figures updated


@pytest.mark.asyncio
async def test_engagement_starts_draft(client: AsyncClient):
    await create_test_company(client, email="dr@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='dr@a.com', password='pass1234')}"}
    resp = await client.post("/api/v1/auditease/engagements", json={"period_label": "FY24"}, headers=headers)
    assert resp.status_code == 201
    assert resp.json()["status"] == EngagementStatus.draft.value


@pytest.mark.asyncio
async def test_engagement_lifecycle(client: AsyncClient):
    await create_test_company(client, email="co@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='co@a.com', password='pass1234')}"}

    await client.post("/api/v1/auth/auditor/register", json={"email": "aud@a.com", "password": "pass1234", "name": "Auditor"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "aud@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    eng_id = await make_engagement(client, co_headers)

    # Invite moves draft -> invited
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "aud@a.com"}, headers=co_headers)
    assert resp.status_code == 200, resp.text
    assert resp.json()["status"] == EngagementStatus.invited.value
    auds = resp.json()["auditors"]
    assert len(auds) == 1
    assert auds[0]["email"] == "aud@a.com"
    assert auds[0]["status"] == GrantStatus.invited.value

    # Auditor sees the invite
    resp = await client.get("/api/v1/auditor/engagements", headers=aud_headers)
    assert len(resp.json()) == 1

    # Accept moves invited -> active
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)
    assert resp.status_code == 200
    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}", headers=co_headers)
    assert resp.json()["status"] == EngagementStatus.active.value

    # Close -> closed, auditor loses access + engagement vanishes from their list
    resp = await client.patch(f"/api/v1/auditease/engagements/{eng_id}/close", headers=co_headers)
    assert resp.status_code == 200
    assert resp.json()["status"] == EngagementStatus.closed.value

    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/trial-balance", headers=aud_headers)
    assert resp.status_code == 403
    resp = await client.get("/api/v1/auditor/engagements", headers=aud_headers)
    assert len(resp.json()) == 0


@pytest.mark.asyncio
async def test_delete_engagement_guard(client: AsyncClient):
    await create_test_company(client, email="del@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='del@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register", json={"email": "deld@a.com", "password": "pass1234", "name": "A"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "deld@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    # Draft engagement can be deleted
    eng_id = await make_engagement(client, co_headers)
    resp = await client.delete(f"/api/v1/auditease/engagements/{eng_id}", headers=co_headers)
    assert resp.status_code == 204

    # Active engagement cannot be deleted
    eng2 = await make_engagement(client, co_headers)
    await client.post(f"/api/v1/auditease/engagements/{eng2}/auditors/invite", json={"email": "deld@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng2}/accept", headers=aud_headers)
    resp = await client.delete(f"/api/v1/auditease/engagements/{eng2}", headers=co_headers)
    assert resp.status_code == 409

    # But once closed, cleanup delete is allowed
    await client.patch(f"/api/v1/auditease/engagements/{eng2}/close", headers=co_headers)
    resp = await client.delete(f"/api/v1/auditease/engagements/{eng2}", headers=co_headers)
    assert resp.status_code == 204


@pytest.mark.asyncio
async def test_pending_invite_autoconverts_on_registration(client: AsyncClient):
    await create_test_company(client, email="pi@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='pi@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, co_headers)

    # Invite an email with no auditor account yet -> pending
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "future@aud.com"}, headers=co_headers)
    assert resp.status_code == 200
    assert resp.json()["status"] == EngagementStatus.invited.value
    assert resp.json()["auditors"][0]["status"] == "pending"

    # Auditor registers with that email -> pending invite becomes a grant
    await client.post("/api/v1/auth/auditor/register", json={"email": "future@aud.com", "password": "pass1234", "name": "Future"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "future@aud.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    resp = await client.get("/api/v1/auditor/engagements", headers=aud_headers)
    assert resp.status_code == 200
    assert len(resp.json()) == 1
    assert resp.json()[0]["id"] == eng_id


async def _engagement_with_entry(client, slug, approve=False):
    """An engagement with an imported TB and one adjusting entry, for re-import tests."""
    await create_test_company(client, email=f"{slug}@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email=f'{slug}@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register",
                      json={"email": f"{slug}aud@a.com", "password": "pass1234", "name": "A"})
    resp = await client.post("/api/v1/auth/auditor/login",
                             json={"email": f"{slug}aud@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    eng_id = await make_engagement(client, co_headers)
    resp = await import_tb(client, eng_id, co_headers)
    ledgers = resp.json()["accounts"]
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite",
                      json={"email": f"{slug}aud@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)

    entry = {
        "description": "Adj",
        "lines": [
            {"ledger_id": ledgers[0]["id"], "side": "debit", "amount": 100},
            {"ledger_id": ledgers[1]["id"], "side": "credit", "amount": 100},
        ],
    }
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/entries",
                             json=entry, headers=aud_headers)
    assert resp.status_code == 201, resp.text
    entry_id = resp.json()["id"]
    if approve:
        resp = await client.patch(
            f"/api/v1/auditease/entries/{entry_id}/approve",
            json={"status": "approved"}, headers=co_headers,
        )
        assert resp.status_code == 200, resp.text
    return eng_id, co_headers, aud_headers, entry_id, ledgers


@pytest.mark.asyncio
async def test_reimport_allowed_with_only_proposed_entries(client: AsyncClient):
    """A merely *proposed* entry must not lock the trial balance.

    The old guard counted any AuditEntry, so one rejected or pending proposal made
    the TB permanently un-reimportable.
    """
    eng_id, co_headers, _, _, _ = await _engagement_with_entry(client, "propose")
    resp = await import_tb(client, eng_id, co_headers)
    assert resp.status_code == 200, resp.text


@pytest.mark.asyncio
async def test_reimport_requires_confirmation_after_approved_entries(client: AsyncClient):
    """Confirm instead of refuse -- and never destroy the approved entry's lines.

    `audit_entry_lines.ledger_id` is ON DELETE CASCADE, so the old delete-then-insert
    import would have silently deleted these lines. That is the regression this pins.
    """
    eng_id, co_headers, _, entry_id, _ = await _engagement_with_entry(client, "approve", approve=True)

    resp = await import_tb(client, eng_id, co_headers)
    assert resp.status_code == 409
    detail = resp.json()["detail"]
    assert detail["reimport_impact"]["approved_entry_count"] == 1
    assert detail["reimport_impact"]["requires_confirmation"] is True

    resp = await import_tb(client, eng_id, co_headers, confirm=True)
    assert resp.status_code == 200, resp.text

    # The approved entry and both of its lines must still exist.
    entries = (await client.get(
        f"/api/v1/auditease/engagements/{eng_id}/entries", headers=co_headers
    )).json()
    kept = next(e for e in entries if e["id"] == entry_id)
    assert kept["status"] == "approved"
    assert len(kept["lines"]) == 2


@pytest.mark.asyncio
async def test_reimport_retains_ledger_referenced_by_entry(client: AsyncClient):
    """A ledger missing from the new file but referenced by an entry line is retained.
    Deleting it would cascade away the adjustment."""
    eng_id, co_headers, _, _, ledgers = await _engagement_with_entry(client, "retain", approve=True)

    # A file that no longer contains "Loan".
    partial = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"A1,Cash,100,50,0,150\n"
    )
    resp = await preview_tb(client, eng_id, co_headers, csv=partial, convention="signed")
    assert resp.status_code == 200, resp.text
    assert "Loan" in resp.json()["reimport_impact"]["retained_referenced"]

    resp = await import_tb(
        client, eng_id, co_headers, csv=partial, convention="signed", confirm=True
    )
    assert resp.status_code == 200, resp.text
    names = [a["ledger_name"] for a in (await get_tb(client, eng_id, co_headers)).json()["accounts"]]
    assert "Loan" in names


async def get_groups(client, headers):
    resp = await client.get("/api/v1/auditease/ledger-groups", headers=headers)
    assert resp.status_code == 200, resp.text
    return resp.json()


def find_group(groups, name):
    return next(g for g in groups if g["name"] == name)


@pytest.mark.asyncio
async def test_chart_of_accounts(client: AsyncClient):
    await create_test_company(client, email="coa2@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='coa2@a.com', password='pass1234')}"}

    groups = await get_groups(client, headers)
    tops = {g["name"] for g in groups if g["level"] == 0}
    assert tops == {"Assets", "Liabilities", "Income", "Expenditure"}
    assets = find_group(groups, "Assets")
    assert assets["company_id"] is None  # seeded, read-only

    # subgroup (level 1)
    resp = await client.post("/api/v1/auditease/ledger-groups", json={"name": "Current Assets", "parent_id": assets["id"]}, headers=headers)
    assert resp.status_code == 201, resp.text
    ca = resp.json()
    assert ca["level"] == 1

    # subsubgroup (level 2)
    resp = await client.post("/api/v1/auditease/ledger-groups", json={"name": "Cash & Bank", "parent_id": ca["id"]}, headers=headers)
    assert resp.status_code == 201
    cb = resp.json()
    assert cb["level"] == 2

    # depth cap
    resp = await client.post("/api/v1/auditease/ledger-groups", json={"name": "Too Deep", "parent_id": cb["id"]}, headers=headers)
    assert resp.status_code == 400

    # cannot rename a seeded top group
    resp = await client.patch(f"/api/v1/auditease/ledger-groups/{assets['id']}", json={"name": "Nope"}, headers=headers)
    assert resp.status_code == 403

    # can rename own group
    resp = await client.patch(f"/api/v1/auditease/ledger-groups/{ca['id']}", json={"name": "Current Assets 2"}, headers=headers)
    assert resp.status_code == 200

    # parent flags updated
    groups = await get_groups(client, headers)
    assert find_group(groups, "Assets")["has_children"] is True
    assert find_group(groups, "Current Assets 2")["has_children"] is True

    # delete guard: has children
    resp = await client.delete(f"/api/v1/auditease/ledger-groups/{ca['id']}", headers=headers)
    assert resp.status_code == 409
    # delete leaf, parent flag clears
    resp = await client.delete(f"/api/v1/auditease/ledger-groups/{cb['id']}", headers=headers)
    assert resp.status_code == 204
    groups = await get_groups(client, headers)
    assert find_group(groups, "Current Assets 2")["has_children"] is False


@pytest.mark.asyncio
async def test_ledger_mapping(client: AsyncClient):
    await create_test_company(client, email="map@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='map@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register", json={"email": "mapaud@a.com", "password": "pass1234", "name": "A"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "mapaud@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    eng_id = await make_engagement(client, co_headers)
    imp = await import_tb(client, eng_id, co_headers)
    ledgers = imp.json()["accounts"]
    cash, loan = ledgers[0]["id"], ledgers[1]["id"]

    groups = await get_groups(client, co_headers)
    assets = find_group(groups, "Assets")
    liab = find_group(groups, "Liabilities")

    # subgroup under Assets, then map cash to the leaf
    resp = await client.post("/api/v1/auditease/ledger-groups", json={"name": "Current Assets", "parent_id": assets["id"]}, headers=co_headers)
    ca = resp.json()

    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/ledgers/{cash}/map", json={"group_id": ca["id"]}, headers=co_headers)
    assert resp.status_code == 200, resp.text
    assert resp.json()["mapped_group_path"] == ["Assets", "Current Assets"]

    # mapping to a non-leaf (Assets now has children) is rejected
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/ledgers/{loan}/map", json={"group_id": assets["id"]}, headers=co_headers)
    assert resp.status_code == 400

    # Liabilities is seeded with Schedule III sub-groups, so it is not a leaf and
    # cannot be mapped to directly.
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/ledgers/{loan}/map", json={"group_id": liab["id"]}, headers=co_headers)
    assert resp.status_code == 400

    # map loan to a company-created Liabilities leaf instead
    resp = await client.post("/api/v1/auditease/ledger-groups", json={"name": "Current Liabilities", "parent_id": liab["id"]}, headers=co_headers)
    cl = resp.json()
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/ledgers/{loan}/map", json={"group_id": cl["id"]}, headers=co_headers)
    assert resp.status_code == 200, resp.text
    assert resp.json()["mapped_group_path"] == ["Liabilities", "Current Liabilities"]

    # can't add a subgroup under a leaf while a ledger is mapped directly to it
    resp = await client.post("/api/v1/auditease/ledger-groups", json={"name": "Provisions", "parent_id": cl["id"]}, headers=co_headers)
    assert resp.status_code == 409

    # company TB reflects the mapping path
    tb = await client.get(f"/api/v1/auditease/engagements/{eng_id}/trial-balance", headers=co_headers)
    cash_row = next(a for a in tb.json()["accounts"] if a["id"] == cash)
    assert cash_row["mapped_group_path"] == ["Assets", "Current Assets"]

    # auditor sees the mapping too
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "mapaud@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)
    tb = await client.get(f"/api/v1/auditor/engagements/{eng_id}/trial-balance", headers=aud_headers)
    cash_row = next(a for a in tb.json()["accounts"] if a["id"] == cash)
    assert cash_row["mapped_group_path"] == ["Assets", "Current Assets"]

    # bulk map then unmap
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/ledgers/bulk-map", json={"ledger_ids": [cash, loan], "group_id": ca["id"]}, headers=co_headers)
    assert resp.json()["updated"] == 2
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/ledgers/unmap", json={"ledger_ids": [cash, loan]}, headers=co_headers)
    assert resp.json()["updated"] == 2
    tb = await client.get(f"/api/v1/auditease/engagements/{eng_id}/trial-balance", headers=co_headers)
    assert all(a["mapped_group_path"] is None for a in tb.json()["accounts"])


@pytest.mark.asyncio
async def test_import_mapping_from_another_engagement(client: AsyncClient):
    await create_test_company(client, email="mapping-import@a.com", password="pass1234")
    headers = {
        "Authorization": f"Bearer {await get_company_token(client, email='mapping-import@a.com', password='pass1234')}"
    }
    source_id = await make_engagement(client, headers, "FY23")
    target_id = await make_engagement(client, headers, "FY24")

    source_csv = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"100,Cash,0,0,0,0\n"
        b"200,Bank,0,0,0,0\n"
        b"DUP,Duplicate,0,0,0,0\n"
        b"DUP,Duplicate,0,0,0,0\n"
        b"AMB,Ambiguous,0,0,0,0\n"
        b"AMB,Ambiguous,0,0,0,0\n"
        b",Trade Receivable,0,0,0,0\n"
    )
    target_csv = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"100,Cash,0,0,0,0\n"
        b"200,Renamed Bank,0,0,0,0\n"
        b"DUP,Duplicate,0,0,0,0\n"
        b"DUP,Duplicate,0,0,0,0\n"
        b"DUP,Duplicate,0,0,0,0\n"
        b"AMB,Ambiguous,0,0,0,0\n"
        b"300,Trade Receivable,0,0,0,0\n"
        b"NO,No Match,0,0,0,0\n"
    )
    source_accounts = (await import_tb(client, source_id, headers, csv=source_csv)).json()["accounts"]
    target_accounts = (await import_tb(client, target_id, headers, csv=target_csv)).json()["accounts"]

    groups = await get_groups(client, headers)
    assets = find_group(groups, "Assets")
    liabilities = find_group(groups, "Liabilities")
    asset_leaf = (
        await client.post(
            "/api/v1/auditease/ledger-groups",
            json={"name": "Mapping Import Assets", "parent_id": assets["id"]},
            headers=headers,
        )
    ).json()
    liability_leaf = (
        await client.post(
            "/api/v1/auditease/ledger-groups",
            json={"name": "Mapping Import Liabilities", "parent_id": liabilities["id"]},
            headers=headers,
        )
    ).json()

    # Duplicate exact identities map one-to-one only. AMB is indistinguishable
    # but points at two groups and therefore must remain unresolved.
    ambiguous_seen = 0
    for account in source_accounts:
        group_id = asset_leaf["id"]
        if account["ledger_code"] == "AMB":
            ambiguous_seen += 1
            if ambiguous_seen == 2:
                group_id = liability_leaf["id"]
        response = await client.post(
            f"/api/v1/auditease/engagements/{source_id}/ledgers/{account['id']}/map",
            json={"group_id": group_id},
            headers=headers,
        )
        assert response.status_code == 200, response.text

    # An incorrect existing target mapping should be overwritten by default.
    response = await client.post(
        f"/api/v1/auditease/engagements/{target_id}/ledgers/"
        f"{next(a for a in target_accounts if a['ledger_code'] == '100')['id']}/map",
        json={"group_id": liability_leaf["id"]},
        headers=headers,
    )
    assert response.status_code == 200, response.text

    sources_response = await client.get(
        f"/api/v1/auditease/engagements/{target_id}/mapping-sources",
        headers=headers,
    )
    assert sources_response.status_code == 200, sources_response.text
    assert sources_response.json() == [{
        "engagement_id": source_id,
        "period_label": "FY23",
        "status": "draft",
        "total_ledger_count": 7,
        "mapped_ledger_count": 7,
    }]

    # One source row is never reused: only two of three DUP targets are assigned.
    # Identifier disagreement and ambiguous source destinations are not guessed.
    response = await client.post(
        f"/api/v1/auditease/engagements/{target_id}/mappings/import",
        json={"source_engagement_id": source_id},
        headers=headers,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert {key: body[key] for key in (
        "total_target_ledgers", "source_mapped_count", "assigned_count",
        "updated_count", "already_correct_count", "preserved_existing_count",
        "unused_source_count", "unresolved_count",
    )} == {
        "total_target_ledgers": 8,
        "source_mapped_count": 7,
        "assigned_count": 4,
        "updated_count": 4,
        "already_correct_count": 0,
        "preserved_existing_count": 0,
        "unused_source_count": 3,
        "unresolved_count": 4,
    }
    assert sorted(
        (item["ledger_code"], item["ledger_name"], item["reason"])
        for item in body["issues"]
    ) == sorted([
        ("200", "Renamed Bank", "identity_disagreement"),
        ("DUP", "Duplicate", "source_exhausted"),
        ("AMB", "Ambiguous", "ambiguous_source_mapping"),
        ("NO", "No Match", "unmatched"),
    ])

    # A repeated overwrite identifies already-correct mappings.
    response = await client.post(
        f"/api/v1/auditease/engagements/{target_id}/mappings/import",
        json={"source_engagement_id": source_id},
        headers=headers,
    )
    assert response.status_code == 200, response.text
    assert response.json()["updated_count"] == 0
    assert response.json()["already_correct_count"] == 4
    assert response.json()["assigned_count"] == 4

    # Preserve mode leaves all four resolved target mappings untouched while
    # still consuming their four source rows.
    response = await client.post(
        f"/api/v1/auditease/engagements/{target_id}/mappings/import",
        json={"source_engagement_id": source_id, "overwrite_existing": False},
        headers=headers,
    )
    assert response.status_code == 200, response.text
    assert response.json()["updated_count"] == 0
    assert response.json()["already_correct_count"] == 0
    assert response.json()["preserved_existing_count"] == 4
    assert response.json()["assigned_count"] == 4

    target_tb = (
        await client.get(
            f"/api/v1/auditease/engagements/{target_id}/trial-balance",
            headers=headers,
        )
    ).json()["accounts"]
    duplicate_targets = [account for account in target_tb if account["ledger_code"] == "DUP"]
    assert sum(account["mapped_group_id"] == asset_leaf["id"] for account in duplicate_targets) == 2
    assert next(account for account in target_tb if account["ledger_code"] == "100")["mapped_group_id"] == asset_leaf["id"]
    assert next(account for account in target_tb if account["ledger_code"] == "200")["mapped_group_id"] is None
    assert next(account for account in target_tb if account["ledger_code"] == "AMB")["mapped_group_id"] is None
    assert next(account for account in target_tb if account["ledger_code"] == "300")["mapped_group_id"] == asset_leaf["id"]
    assert next(account for account in target_tb if account["ledger_code"] == "NO")["mapped_group_id"] is None


@pytest.mark.asyncio
async def test_mapping_import_validation_and_tenant_isolation(client: AsyncClient):
    await create_test_company(client, email="mapping-a@a.com", password="pass1234")
    headers_a = {
        "Authorization": f"Bearer {await get_company_token(client, email='mapping-a@a.com', password='pass1234')}"
    }
    source_id = await make_engagement(client, headers_a, "Source")
    target_without_tb = await make_engagement(client, headers_a, "Target")

    response = await client.post(
        f"/api/v1/auditease/engagements/{target_without_tb}/mappings/import",
        json={"source_engagement_id": target_without_tb},
        headers=headers_a,
    )
    assert response.status_code == 400

    await import_tb(client, source_id, headers_a)
    await import_tb(client, target_without_tb, headers_a)
    response = await client.post(
        f"/api/v1/auditease/engagements/{target_without_tb}/mappings/import",
        json={"source_engagement_id": source_id},
        headers=headers_a,
    )
    assert response.status_code == 409
    assert "no mapped ledgers" in response.json()["detail"]

    await create_test_company(client, email="mapping-b@a.com", password="pass1234")
    headers_b = {
        "Authorization": f"Bearer {await get_company_token(client, email='mapping-b@a.com', password='pass1234')}"
    }
    foreign_target = await make_engagement(client, headers_b, "Foreign")
    await import_tb(client, foreign_target, headers_b)

    response = await client.post(
        f"/api/v1/auditease/engagements/{foreign_target}/mappings/import",
        json={"source_engagement_id": source_id},
        headers=headers_b,
    )
    assert response.status_code == 404


@pytest.mark.asyncio
async def test_audit_entries(client: AsyncClient):
    await create_test_company(client, email="co2@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='co2@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register", json={"email": "aud2@a.com", "password": "pass1234", "name": "Auditor"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "aud2@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    eng_id = await make_engagement(client, co_headers)
    resp = await import_tb(client, eng_id, co_headers)
    ledgers = resp.json()["accounts"]
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "aud2@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)

    entry_data = {
        "description": "Adjusting entry",
        "lines": [
            {"ledger_id": ledgers[0]["id"], "side": "debit", "amount": 100},
            {"ledger_id": ledgers[1]["id"], "side": "credit", "amount": 100},
        ],
    }
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/entries", json=entry_data, headers=aud_headers)
    assert resp.status_code == 201, resp.text
    entry_id = resp.json()["id"]

    resp = await client.patch(f"/api/v1/auditease/entries/{entry_id}/approve", json={"status": "approved"}, headers=co_headers)
    assert resp.status_code == 200
    assert resp.json()["status"] == "approved"


@pytest.mark.asyncio
async def test_requirements_and_queries(client: AsyncClient):
    await create_test_company(client, email="co3@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='co3@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register", json={"email": "aud3@a.com", "password": "pass1234", "name": "Auditor"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "aud3@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    eng_id = await make_engagement(client, co_headers)
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "aud3@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)

    # create with defaults -> REQ-001 open P1
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests",
                             json={"description": "Provide bank statements"}, headers=aud_headers)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["status"] == "open"
    assert body["requirement_id_str"] == "REQ-001"
    assert body["priority"] == 1
    req_id = body["id"]

    # second requirement gets REQ-002; priority honored
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests",
                             json={"description": "Ledger dump", "priority": 3}, headers=aud_headers)
    assert resp.json()["requirement_id_str"] == "REQ-002"
    second_id = resp.json()["id"]

    # respond with text only -> status stays open, submission_count == 1
    resp = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/requirement-requests/{req_id}/respond",
        data={"text_answer": "Will upload Monday"},
        headers=co_headers,
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["status"] == "open"
    assert resp.json()["submission_count"] == 1
    assert len(resp.json()["submissions"]) == 1
    assert resp.json()["submissions"][0]["round_number"] == 1

    # second response -> round_number == 2, two submissions, still open
    resp = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/requirement-requests/{req_id}/respond",
        data={"text_answer": "Here is additional context"},
        headers=co_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "open"
    assert resp.json()["submission_count"] == 2
    assert len(resp.json()["submissions"]) == 2
    assert resp.json()["submissions"][1]["round_number"] == 2

    # auditor closes requirement
    resp = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/{req_id}/close",
        headers=aud_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "closed"
    assert resp.json()["closed_at"] is not None

    # respond while closed -> 400
    resp = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/requirement-requests/{req_id}/respond",
        data={"text_answer": "late"},
        headers=co_headers,
    )
    assert resp.status_code == 400

    # edit while closed -> 400
    resp = await client.put(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/{req_id}",
        json={"description": "edited"},
        headers=aud_headers,
    )
    assert resp.status_code == 400

    # auditor reopens requirement
    resp = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/{req_id}/reopen",
        headers=aud_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "open"
    assert resp.json()["closed_at"] is None

    # edit now succeeds
    resp = await client.put(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/{req_id}",
        json={"description": "Provide bank statements for all accounts", "priority": 2},
        headers=aud_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["description"] == "Provide bank statements for all accounts"
    assert resp.json()["priority"] == 2

    # delete with submissions -> 400
    resp = await client.delete(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/{req_id}",
        headers=aud_headers,
    )
    assert resp.status_code == 400

    # empty respond -> 422
    resp = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/requirement-requests/{second_id}/respond",
        data={},
        headers=co_headers,
    )
    assert resp.status_code == 422

    # delete without submissions succeeds
    resp = await client.delete(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/{second_id}",
        headers=aud_headers,
    )
    assert resp.status_code == 200

    # queries link to requirements
    resp = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/queries",
        data={"initial_message": "What is this?", "requirement_id": req_id},
        headers=aud_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["requirement_id"] == req_id


@pytest.mark.asyncio
async def test_auditease_cross_tenant_leak(client: AsyncClient):
    await create_test_company(client, email="coa@a.com", password="pass1234")
    headers_a = {"Authorization": f"Bearer {await get_company_token(client, email='coa@a.com', password='pass1234')}"}
    await create_test_company(client, email="cob@a.com", password="pass1234")
    headers_b = {"Authorization": f"Bearer {await get_company_token(client, email='cob@a.com', password='pass1234')}"}

    eng_id = await make_engagement(client, headers_a)
    await import_tb(client, eng_id, headers_a)

    # B cannot read A's engagement TB
    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}/trial-balance", headers=headers_b)
    assert resp.status_code == 404

    # B cannot see A's engagements
    resp = await client.get("/api/v1/auditease/engagements", headers=headers_b)
    assert len(resp.json()) == 0

    # B cannot close A's engagement
    resp = await client.patch(f"/api/v1/auditease/engagements/{eng_id}/close", headers=headers_b)
    assert resp.status_code == 404

    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/requirement-requests/{uuid.uuid4()}/respond",
                             data={"text_answer": "hi"}, headers=headers_b)
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_auditor_document_access_and_queries(client: AsyncClient):
    await create_test_company(client, email="co4@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='co4@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register", json={"email": "aud4@a.com", "password": "pass1234", "name": "Auditor"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "aud4@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}
    
    # second auditor for cross-check
    await client.post("/api/v1/auth/auditor/register", json={"email": "aud_other@a.com", "password": "pass1234", "name": "Other"})
    resp2 = await client.post("/api/v1/auth/auditor/login", json={"email": "aud_other@a.com", "password": "pass1234"})
    aud_other_headers = {"Authorization": f"Bearer {resp2.json()['access_token']}"}

    eng_id = await make_engagement(client, co_headers)
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "aud4@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)
    
    # Auditor raises a query with a file
    files = {'file': ('query_doc.txt', b'query content', 'text/plain')}
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/queries", data={"initial_message": "Query 1"}, files=files, headers=aud_headers)
    assert resp.status_code == 200
    query_id = resp.json()["id"]
    q_msg = resp.json()["messages"][0]
    q_doc_id = q_msg["attached_document_id"]
    assert q_doc_id is not None
    
    # Auditor can download it
    resp = await client.get(f"/api/v1/auditor/documents/{q_doc_id}/download", headers=aud_headers)
    assert resp.status_code == 200
    assert resp.content == b'query content'
    
    # Other auditor cannot access it
    resp = await client.get(f"/api/v1/auditor/documents/{q_doc_id}/download", headers=aud_other_headers)
    assert resp.status_code == 404
    
    # Company responds with file
    c_files = {'file': ('reply_doc.txt', b'reply content', 'text/plain')}
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/queries/{query_id}/messages", data={"text": "Here is reply"}, files=c_files, headers=co_headers)
    assert resp.status_code == 200
    c_doc_id = resp.json()["attached_document_id"]
    
    # Auditor can download reply file
    resp = await client.get(f"/api/v1/auditor/documents/{c_doc_id}/download", headers=aud_headers)
    assert resp.status_code == 200
    assert resp.content == b'reply content'
    
    # Auditor lists queries
    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/queries", headers=aud_headers)
    assert resp.status_code == 200
    assert len(resp.json()) == 1
    
    # Auditor closes query
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/queries/{query_id}/close", headers=aud_headers)
    assert resp.status_code == 200
    
    # Can no longer add messages
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/queries/{query_id}/messages", data={"text": "late"}, headers=aud_headers)
    assert resp.status_code == 400
    
    # Company closes engagement
    await client.patch(f"/api/v1/auditease/engagements/{eng_id}/close", headers=co_headers)
    
    # Auditor can no longer download documents
    resp = await client.get(f"/api/v1/auditor/documents/{c_doc_id}/download", headers=aud_headers)
    assert resp.status_code == 404


# --- Entry ledger names + report preview ---------------------------------------

async def _accept_auditor(client, co_headers, eng_id, email):
    await client.post("/api/v1/auth/auditor/register", json={"email": email, "password": "pass1234", "name": "A"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": email, "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": email}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)
    return aud_headers


@pytest.mark.asyncio
async def test_entry_lines_include_ledger_name(client: AsyncClient):
    """Both the auditor and company entry views must carry the ledger name/code so
    the UI never shows 'Unknown Ledger'."""
    await create_test_company(client, email="eln@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='eln@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, co_headers)
    imp = await import_tb(client, eng_id, co_headers)
    ledgers = {a["ledger_name"]: a for a in imp.json()["accounts"]}
    cash, loan = ledgers["Cash"], ledgers["Loan"]
    aud_headers = await _accept_auditor(client, co_headers, eng_id, "elnaud@a.com")

    entry = {
        "code": "AJE-1",
        "description": "Reclass",
        "lines": [
            {"ledger_id": cash["id"], "side": "debit", "amount": 100},
            {"ledger_id": loan["id"], "side": "credit", "amount": 100},
        ],
    }
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/entries", json=entry, headers=aud_headers)
    assert resp.status_code == 201, resp.text
    # The create response already carries ledger identity.
    created_lines = {l["ledger_id"]: l for l in resp.json()["lines"]}
    assert created_lines[cash["id"]]["ledger_name"] == "Cash"
    assert created_lines[cash["id"]]["ledger_code"] == "A1"
    entry_id = resp.json()["id"]

    # Auditor list view
    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/entries", headers=aud_headers)
    aud_lines = {l["ledger_id"]: l for l in resp.json()[0]["lines"]}
    assert aud_lines[cash["id"]]["ledger_name"] == "Cash"
    assert aud_lines[loan["id"]]["ledger_name"] == "Loan"

    # Company list view
    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}/entries", headers=co_headers)
    co_lines = {l["ledger_id"]: l for l in resp.json()[0]["lines"]}
    assert co_lines[cash["id"]]["ledger_name"] == "Cash"
    assert co_lines[loan["id"]]["ledger_name"] == "Loan"

    # Approve response also carries ledger identity
    resp = await client.patch(f"/api/v1/auditease/entries/{entry_id}/approve", json={"status": "approved"}, headers=co_headers)
    assert resp.status_code == 200
    ap_lines = {l["ledger_id"]: l for l in resp.json()["lines"]}
    assert ap_lines[loan["id"]]["ledger_name"] == "Loan"


@pytest.mark.asyncio
async def test_delete_closed_engagement_with_children(client: AsyncClient):
    """A closed engagement that accumulated entries, a query and a requirement must
    still delete cleanly (regression: FK 500 when children weren't cascaded)."""
    await create_test_company(client, email="delc@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='delc@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, co_headers)
    imp = await import_tb(client, eng_id, co_headers)
    ledgers = imp.json()["accounts"]
    aud_headers = await _accept_auditor(client, co_headers, eng_id, "delcaud@a.com")

    # Auditor adds an entry, a requirement and a query (the previously-uncascaded rows).
    await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/entries",
        json={"description": "Adj", "lines": [
            {"ledger_id": ledgers[0]["id"], "side": "debit", "amount": 100},
            {"ledger_id": ledgers[1]["id"], "side": "credit", "amount": 100},
        ]},
        headers=aud_headers,
    )
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests", json={"description": "docs"}, headers=aud_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/queries", data={"initial_message": "hi"}, headers=aud_headers)

    # Close, then delete — must succeed, not 500.
    await client.patch(f"/api/v1/auditease/engagements/{eng_id}/close", headers=co_headers)
    resp = await client.delete(f"/api/v1/auditease/engagements/{eng_id}", headers=co_headers)
    assert resp.status_code == 204, resp.text

    # Gone.
    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}", headers=co_headers)
    assert resp.status_code == 404


REPORT_CSV = (
    b"Code,Name,Opening,Debit,Credit,Closing\n"
    b"A1,Cash,0,0,0,1000\n"
    b"L1,Loan,0,0,0,600\n"
    b"I1,Sales,0,0,0,500\n"
    b"E1,Rent,0,0,0,100\n"
    b"U1,Suspense,0,0,0,999\n"
)


# Same figures as REPORT_CSV but in the standard SIGNED trial-balance convention:
# credit-natured accounts (Income, Liabilities) carry a negative closing balance.
SIGNED_REPORT_CSV = (
    b"Code,Name,Opening,Debit,Credit,Closing\n"
    b"A1,Cash,0,0,0,1000\n"
    b"L1,Loan,0,0,600,-600\n"
    b"I1,Sales,0,0,500,-500\n"
    b"E1,Rent,0,100,0,100\n"
)

# Signed convention where expenditure exceeds income -> a genuine net LOSS.
SIGNED_LOSS_CSV = (
    b"Code,Name,Opening,Debit,Credit,Closing\n"
    b"A1,Cash,0,0,0,1000\n"
    b"L1,Loan,0,0,900,-900\n"
    b"I1,Sales,0,0,100,-100\n"
    b"E1,Rent,0,500,0,500\n"
)


async def _map_pl_ledgers(client, eng_id, co_headers, ledgers):
    """Map Cash/Loan/Sales/Rent to the seeded Schedule III leaves (shared setup)."""
    groups = await get_groups(client, co_headers)

    async def map_to(ledger_name, group_name):
        gid = find_group(groups, group_name)["id"]
        lid = ledgers[ledger_name]["id"]
        r = await client.post(
            f"/api/v1/auditease/engagements/{eng_id}/ledgers/{lid}/map",
            json={"group_id": gid}, headers=co_headers,
        )
        assert r.status_code == 200, r.text

    await map_to("Cash", "Cash and Cash Equivalents")
    await map_to("Loan", "Trade Payables")
    await map_to("Sales", "Revenue from Operations")
    await map_to("Rent", "Other Expenses")


@pytest.mark.asyncio
async def test_report_preview_signed_trial_balance(client: AsyncClient):
    """Regression: a signed trial balance (credit accounts negative) must still yield
    net profit = Income - Expenditure, not -(|Income| + |Expenditure|). Previously the
    negative Income flipped the subtraction into an addition and reported a false loss."""
    await create_test_company(client, email="signed@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='signed@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, co_headers)
    imp = await import_tb(client, eng_id, co_headers, csv=SIGNED_REPORT_CSV)
    ledgers = {a["ledger_name"]: a for a in imp.json()["accounts"]}
    await _map_pl_ledgers(client, eng_id, co_headers, ledgers)

    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}/reports/preview", headers=co_headers)
    assert resp.status_code == 200, resp.text
    p = resp.json()
    # Totals are reported as positive natural-side magnitudes despite the negative source.
    assert {k: p["totals"][k] for k in ("assets", "liabilities", "income", "expenditure")} == {
        "assets": 1000.0, "liabilities": 600.0,
        "income": 500.0, "expenditure": 100.0,
    }
    # The net is the difference (500 - 100 = 400 profit) — NOT the -600 sum-of-magnitudes.
    assert p["net_profit"] == 400.0
    # And the balance sheet reconciles (Liabilities is also normalized to a magnitude).
    assert p["balance_check"]["liabilities_plus_equity"] == 1000.0
    assert p["balance_check"]["balanced"] is True


@pytest.mark.asyncio
async def test_report_preview_signed_net_loss(client: AsyncClient):
    """A signed trial balance where Expenditure > Income reports a real net LOSS
    (negative net_profit), confirming the sign is genuine and not a flip artifact."""
    await create_test_company(client, email="loss@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='loss@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, co_headers)
    imp = await import_tb(client, eng_id, co_headers, csv=SIGNED_LOSS_CSV)
    ledgers = {a["ledger_name"]: a for a in imp.json()["accounts"]}
    await _map_pl_ledgers(client, eng_id, co_headers, ledgers)

    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}/reports/preview", headers=co_headers)
    assert resp.status_code == 200, resp.text
    p = resp.json()
    assert p["totals"]["income"] == 100.0
    assert p["totals"]["expenditure"] == 500.0
    # Income 100 - Expenditure 500 = -400 (a real loss), not +600 or -600.
    assert p["net_profit"] == -400.0


@pytest.mark.asyncio
async def test_report_contra_balances_reduce_credit_nature_groups(client: AsyncClient):
    await create_test_company(client, email="contra@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='contra@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)
    csv = (
        b"Code,Name,Opening,Debit,Credit,Closing\n"
        b"A1,Cash,0,0,0,1000\n"
        b"L1,Loan,0,0,0,-600\n"
        b"R1,Accumulated Deficit,0,0,0,250\n"
        b"I1,Sales,0,0,0,-850\n"
        b"I2,Sales Returns,0,0,0,100\n"
        b"E1,Rent,0,0,0,100\n"
    )
    response = await import_tb(client, eng_id, headers, csv=csv, convention="signed")
    assert response.status_code == 200, response.text
    ledgers = {a["ledger_name"]: a for a in response.json()["accounts"]}
    groups = await get_groups(client, headers)
    destinations = {
        "Cash": "Cash and Cash Equivalents",
        "Loan": "Trade Payables",
        "Accumulated Deficit": "Reserves & Surplus",
        "Sales": "Revenue from Operations",
        "Sales Returns": "Revenue from Operations",
        "Rent": "Other Expenses",
    }
    for ledger_name, group_name in destinations.items():
        result = await client.post(
            f"/api/v1/auditease/engagements/{eng_id}/ledgers/{ledgers[ledger_name]['id']}/map",
            json={"group_id": find_group(groups, group_name)["id"]}, headers=headers,
        )
        assert result.status_code == 200, result.text

    report = (await client.get(
        f"/api/v1/auditease/engagements/{eng_id}/reports/preview", headers=headers
    )).json()
    assert report["totals"]["liabilities"] == 350.0
    assert report["totals"]["income"] == 750.0
    assert report["net_profit"] == 650.0
    assert report["balance_check"]["balanced"] is True
    assert report["balance_check"]["statement_ready"] is True


@pytest.mark.asyncio
async def test_tb_sign_convention_repair_with_approved_entry(client: AsyncClient):
    await create_test_company(client, email="repair@a.com", password="pass1234")
    headers = {"Authorization": f"Bearer {await get_company_token(client, email='repair@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, headers)
    csv = b"Code,Name,Closing\nA1,Cash,100\nL1,Loan,100\n"
    cmap = {"ledger_code": "Code", "ledger_name": "Name", "closing_balance": "Closing"}
    imported = await import_tb(client, eng_id, headers, csv=csv, cmap=cmap, convention="signed")
    ledgers = {a["ledger_name"]: a for a in imported.json()["accounts"]}
    groups = await get_groups(client, headers)
    for name, destination in (("Cash", "Cash and Cash Equivalents"), ("Loan", "Trade Payables")):
        await client.post(
            f"/api/v1/auditease/engagements/{eng_id}/ledgers/{ledgers[name]['id']}/map",
            json={"group_id": find_group(groups, destination)["id"]}, headers=headers,
        )
    assert (await get_tb(client, eng_id, headers)).json()["totals"]["balanced"] is False

    auditor_headers = await _accept_auditor(client, headers, eng_id, "repairaud@a.com")
    entry = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/entries",
        json={"description": "Keep me", "lines": [
            {"ledger_id": ledgers["Cash"]["id"], "side": "debit", "amount": 10},
            {"ledger_id": ledgers["Loan"]["id"], "side": "credit", "amount": 10},
        ]}, headers=auditor_headers,
    )
    await client.patch(
        f"/api/v1/auditease/entries/{entry.json()['id']}/approve",
        json={"status": "approved"}, headers=headers,
    )
    repaired = await client.post(
        f"/api/v1/auditease/engagements/{eng_id}/trial-balance/sign-convention",
        json={"convention": "magnitude"}, headers=headers,
    )
    assert repaired.status_code == 200, repaired.text
    assert repaired.json()["totals"]["balanced"] is True
    assert repaired.json()["sign_convention"] == "magnitude"


@pytest.mark.asyncio
async def test_report_preview(client: AsyncClient):
    await create_test_company(client, email="rep@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='rep@a.com', password='pass1234')}"}
    eng_id = await make_engagement(client, co_headers)
    imp = await import_tb(client, eng_id, co_headers, csv=REPORT_CSV)
    ledgers = {a["ledger_name"]: a for a in imp.json()["accounts"]}

    groups = await get_groups(client, co_headers)

    async def map_to(ledger_name, group_name):
        gid = find_group(groups, group_name)["id"]
        lid = ledgers[ledger_name]["id"]
        r = await client.post(
            f"/api/v1/auditease/engagements/{eng_id}/ledgers/{lid}/map",
            json={"group_id": gid}, headers=co_headers,
        )
        assert r.status_code == 200, r.text

    # Map four ledgers to seeded Schedule III leaves; leave Suspense unmapped.
    await map_to("Cash", "Cash and Cash Equivalents")
    await map_to("Loan", "Trade Payables")
    await map_to("Sales", "Revenue from Operations")
    await map_to("Rent", "Other Expenses")

    # --- Preview before any entries -------------------------------------------
    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}/reports/preview", headers=co_headers)
    assert resp.status_code == 200, resp.text
    p = resp.json()
    assert {k: p["totals"][k] for k in ("assets", "liabilities", "income", "expenditure")} == {
        "assets": 1000.0, "liabilities": 600.0,
        "income": 500.0, "expenditure": 100.0,
    }
    assert p["net_profit"] == 400.0
    assert p["balance_check"]["liabilities_plus_equity"] == 1000.0
    assert p["balance_check"]["balanced"] is True
    assert p["unmapped_count"] == 1
    assert p["entries"]["approved_count"] == 0
    assert p["entries"]["proposed_count"] == 0

    # --- Add + approve an adjusting entry -------------------------------------
    aud_headers = await _accept_auditor(client, co_headers, eng_id, "repaud@a.com")
    entry = {
        "description": "Extra sale",
        "lines": [
            {"ledger_id": ledgers["Cash"]["id"], "side": "debit", "amount": 200},
            {"ledger_id": ledgers["Sales"]["id"], "side": "credit", "amount": 200},
        ],
    }
    resp = await client.post(f"/api/v1/auditor/engagements/{eng_id}/entries", json=entry, headers=aud_headers)
    entry_id = resp.json()["id"]
    await client.patch(f"/api/v1/auditease/entries/{entry_id}/approve", json={"status": "approved"}, headers=co_headers)

    # Second, un-approved entry -> counted as proposed only
    entry2 = {
        "description": "Pending",
        "lines": [
            {"ledger_id": ledgers["Rent"]["id"], "side": "debit", "amount": 50},
            {"ledger_id": ledgers["Loan"]["id"], "side": "credit", "amount": 50},
        ],
    }
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/entries", json=entry2, headers=aud_headers)

    resp = await client.get(f"/api/v1/auditease/engagements/{eng_id}/reports/preview", headers=co_headers)
    p = resp.json()
    # Approved debit to Cash raises assets; credit to Sales raises income.
    assert p["totals"]["assets"] == 1200.0
    assert p["totals"]["income"] == 700.0
    assert p["net_profit"] == 600.0
    assert p["balance_check"]["balanced"] is True  # double-entry keeps it balanced
    assert p["entries"]["approved_count"] == 1
    assert p["entries"]["proposed_count"] == 1

    tb_view = (await get_tb(client, eng_id, co_headers)).json()
    assert tb_view["totals"]["difference"] == p["balance_check"]["difference"]
    assert tb_view["totals"]["balanced"] == p["balance_check"]["balanced"]
    assert tb_view["totals"]["assets"] == p["totals"]["assets"]
    assert tb_view["totals"]["income"] == p["totals"]["income"]
    cash_view = next(a for a in tb_view["accounts"] if a["ledger_name"] == "Cash")
    assert cash_view["adjustment_net_debit"] == 200.0
    assert cash_view["final_net_debit"] == 1200.0

    cash_line = next(l for l in p["lines"] if l["ledger_name"] == "Cash")
    assert cash_line["adjustment"] == 200.0
    assert cash_line["final"] == 1200.0
    assert cash_line["top_group"] == "Assets"

    # --- Generate persists an HTML report to docVault --------------------------
    resp = await client.post(f"/api/v1/auditease/engagements/{eng_id}/reports/generate", headers=co_headers)
    assert resp.status_code == 200, resp.text
    assert "id" in resp.json() and "url" in resp.json()


# --- Report HTML rendering (Profit & Loss Schedule III) -----------------------

import uuid as _uuid
from decimal import Decimal

from app.models.auditease import BalanceNature
from app.services.reporting.auditease_reports import build_profit_and_loss
from app.services.reporting.pdf import render_html
from app.services.trial_balance import LedgerFigure, TBSummary, summarize


def _make_figures(*, income: Decimal, expenditure: Decimal):
    sales_fig = LedgerFigure(
        ledger_id=_uuid.uuid4(),
        ledger_name="Sales",
        ledger_code="I1",
        top_group="Income",
        group_path=["Income", "Revenue from Operations"],
        nature=BalanceNature.credit,
        opening_net_debit=Decimal(0),
        net_debit=-income,
        adjustment=Decimal(0),
        final_net_debit=-income,
        presented_closing=income,
        presented_final=income,
    )
    rent_fig = LedgerFigure(
        ledger_id=_uuid.uuid4(),
        ledger_name="Rent",
        ledger_code="E1",
        top_group="Expenditure",
        group_path=["Expenditure", "Other Expenses"],
        nature=BalanceNature.debit,
        opening_net_debit=Decimal(0),
        net_debit=expenditure,
        adjustment=Decimal(0),
        final_net_debit=expenditure,
        presented_closing=expenditure,
        presented_final=expenditure,
    )
    figs = [sales_fig, rent_fig]
    summary = summarize(figs)
    return figs, summary


def test_report_html_pl_net_is_difference_not_sum():
    """Regression: the P&L bottom line must be Income - Expenditure, and not a sum."""
    figs, summary = _make_figures(income=Decimal("700.00"), expenditure=Decimal("100.00"))
    doc = build_profit_and_loss(figs, summary, "Test Co", "FY24", "absolute")
    html = render_html(doc)

    # Total revenue is 700.00, total expenses 100.00, profit before tax is 600.00
    assert "700.00" in html
    assert "100.00" in html
    assert "600.00" in html
    # The misleading sum (700 + 100 = 800) must never surface as the net.
    assert "800.00" not in html


def test_report_html_pl_reports_a_loss():
    """When Expenditure exceeds Income the report shows a Net Loss of the difference."""
    figs, summary = _make_figures(income=Decimal("100.00"), expenditure=Decimal("700.00"))
    doc = build_profit_and_loss(figs, summary, "Test Co", "FY24", "absolute")
    html = render_html(doc)

    # Net loss is difference (600.00), not sum (800.00)
    assert "600.00" in html
    assert "800.00" not in html


@pytest.mark.asyncio
async def test_requirement_bulk_import_roundtrip(client: AsyncClient):
    import io
    import openpyxl
    from app.services.requirement_import import build_template_xlsx

    await create_test_company(client, email="coi@a.com", password="pass1234")
    co_headers = {"Authorization": f"Bearer {await get_company_token(client, email='coi@a.com', password='pass1234')}"}
    await client.post("/api/v1/auth/auditor/register", json={"email": "audi@a.com", "password": "pass1234", "name": "Aud"})
    resp = await client.post("/api/v1/auth/auditor/login", json={"email": "audi@a.com", "password": "pass1234"})
    aud_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}
    eng_id = await make_engagement(client, co_headers)
    await client.post(f"/api/v1/auditease/engagements/{eng_id}/auditors/invite", json={"email": "audi@a.com"}, headers=co_headers)
    await client.post(f"/api/v1/auditor/engagements/{eng_id}/accept", headers=aud_headers)

    # template downloads
    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/import-template",
                            headers=aud_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # build a filled sheet: 2 rows with 4 columns: S. No., Requirement, Due Date, Priority
    wb = openpyxl.load_workbook(io.BytesIO(build_template_xlsx()))
    ws = wb["Requirements"]
    ws.delete_rows(2)  # drop example row
    ws.append(["1", "Bulk req A", None, None])
    ws.append(["2", "Bulk req B", "2026-09-15", 4])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    resp = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/import",
        files={"file": ("reqs.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=aud_headers)
    assert resp.status_code == 200, resp.text
    assert resp.json()["created_count"] == 2

    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests", headers=aud_headers)
    by_desc = {r["description"]: r for r in resp.json()}
    assert by_desc["Bulk req A"]["requirement_id_str"] == "REQ-001"
    assert by_desc["Bulk req A"]["priority"] == 1
    assert by_desc["Bulk req B"]["requirement_id_str"] == "REQ-002"
    assert by_desc["Bulk req B"]["priority"] == 4

    # second import appends REQ-003+
    wb_second = openpyxl.load_workbook(io.BytesIO(build_template_xlsx()))
    ws_sec = wb_second["Requirements"]
    ws_sec.delete_rows(2)
    ws_sec.append(["1", "Bulk req C", None, 2])
    buf_sec = io.BytesIO(); wb_second.save(buf_sec); buf_sec.seek(0)
    resp = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/import",
        files={"file": ("reqs_sec.xlsx", buf_sec.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=aud_headers)
    assert resp.status_code == 200
    assert resp.json()["created_count"] == 1

    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests", headers=aud_headers)
    by_desc = {r["description"]: r for r in resp.json()}
    assert by_desc["Bulk req C"]["requirement_id_str"] == "REQ-003"

    # all-or-nothing: one bad row aborts everything (bad date)
    wb2 = openpyxl.load_workbook(io.BytesIO(build_template_xlsx()))
    ws2 = wb2["Requirements"]
    ws2.delete_rows(2)
    ws2.append(["1", "Good row", None, None])
    ws2.append(["2", "Bad row", "not-a-date", None])
    buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
    resp = await client.post(
        f"/api/v1/auditor/engagements/{eng_id}/requirement-requests/import",
        files={"file": ("bad.xlsx", buf2.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=aud_headers)
    assert resp.status_code == 422
    assert any(e["row"] == 3 for e in resp.json()["detail"])

    resp = await client.get(f"/api/v1/auditor/engagements/{eng_id}/requirement-requests", headers=aud_headers)
    assert len(resp.json()) == 3  # nothing extra persisted
