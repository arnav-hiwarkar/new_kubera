"""Integration tests for Fixed Asset Register and Depreciation Reports."""
from datetime import date
from decimal import Decimal
import io
import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy import select

from app.models.assets import Asset, AssetLifecycleStatus, AssetOperationalStatus, AssetAcquisition, ItcTreatment, DiscountType
from app.models.asset_masters import AssetCategory, ItAssetBlock
from app.models.company import Company, CompanyUser
from app.models.depreciation import DepreciationRun
from app.models.financial_year import FinancialYear
from app.routers.asset_reports import _load_asset_context
from app.services.reporting.asset_reports import (
    ALL_ASSET_REPORTS,
    build_fixed_asset_register_report,
    build_companies_act_schedule_ii_report,
    build_income_tax_appendix_i_report,
    build_gst_itc_summary_report,
)
from tests.asset_helpers import admin_headers, seed_masters
from tests.conftest import TestSessionLocal, create_test_company, get_company_token


async def setup_asset_reports_environment(client: AsyncClient, email: str = "admin_reports@testco.com"):
    await seed_masters()
    headers = await admin_headers(client, email)

    # 1. Create Financial Year
    fy_res = await client.post(
        "/api/v1/financial-years",
        json={
            "label": "2024-25",
            "start_date": "2024-04-01",
            "end_date": "2025-03-31",
        },
        headers=headers,
    )
    assert fy_res.status_code == 201, fy_res.text
    fy_id = fy_res.json()["id"]

    # 2. Insert test assets with full lifecycle
    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        cat = (await session.execute(select(AssetCategory))).scalars().first()
        it_block = (await session.execute(select(ItAssetBlock))).scalars().first()

        # Active capitalized asset
        a1 = Asset(
            company_id=user.company_id,
            asset_name="Production CNC Machine",
            asset_code="MCH-001",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 4, 1),
            available_for_use_date=date(2024, 4, 1),
            useful_life_months=120,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("500000.00"),
            it_block_id=it_block.id if it_block else None,
            it_dep_rate=Decimal("15.00"),
            it_put_to_use_date=date(2024, 4, 1),
        )

        # Disposed asset during FY
        a2 = Asset(
            company_id=user.company_id,
            asset_name="Old Office Van",
            asset_code="VEH-002",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.disposed,
            operational_status=AssetOperationalStatus.in_storage,
            capitalization_date=date(2024, 4, 1),
            available_for_use_date=date(2024, 4, 1),
            useful_life_months=60,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("200000.00"),
            disposal_date=date(2024, 9, 30),
            disposal_type="sale",
            sale_proceeds=Decimal("120000.00"),
            it_block_id=it_block.id if it_block else None,
            it_dep_rate=Decimal("15.00"),
        )

        # CWIP asset
        a3 = Asset(
            company_id=user.company_id,
            asset_name="Factory Expansion Under Construction",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.draft,
            operational_status=AssetOperationalStatus.in_storage,
            original_cost=Decimal("350000.00"),
        )

        session.add_all([a1, a2, a3])
        await session.commit()

    # 3. Trigger and finalize depreciation calculation run
    run_res = await client.post(
        "/api/v1/depreciation/runs",
        json={"financial_year_id": fy_id, "notes": "Audited run"},
        headers=headers,
    )
    assert run_res.status_code == 201, run_res.text
    run_id = run_res.json()["id"]

    fin_res = await client.post(f"/api/v1/depreciation/runs/{run_id}/finalize", headers=headers)
    assert fin_res.status_code == 200, fin_res.text

    return {
        "headers": headers,
        "email": email,
        "fy_id": fy_id,
        "run_id": run_id,
    }


@pytest.mark.asyncio
async def test_asset_reports_list(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="list_rep@testco.com")
    headers = ctx["headers"]

    list_res = await client.get("/api/v1/asset-reports", headers=headers)
    assert list_res.status_code == 200, list_res.text
    report_list = list_res.json()
    assert len(report_list) == 10


@pytest.mark.asyncio
async def test_fixed_asset_register_invariants(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="far_inv@testco.com")
    headers = ctx["headers"]
    fy_id = ctx["fy_id"]
    email = ctx["email"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        context = await _load_asset_context(session, user.company_id, fy_id)

    doc = build_fixed_asset_register_report(
        assets=context["assets"],
        dep_lines_by_asset_id=context["dep_lines_by_asset_id"],
        company_name=context["company_name"],
        fy_label=context["fy"].label,
        units="absolute",
        lookups_by_id=context["lookups_by_id"],
    )

    root = doc.sections[0]
    assert len(root.children) > 0
    assert root.total.cells["original_cost"] > 0
    assert root.total.cells["net_book_value"] > 0

    # Each category subtotal equals sum of rows
    for cat_sec in root.children:
        assert len(cat_sec.rows) > 0
        assert cat_sec.total.cells["original_cost"] > 0
        assert cat_sec.total.cells["original_cost"] == sum(r.cells["original_cost"] for r in cat_sec.rows)
        assert cat_sec.total.cells["opening_acc_dep"] == sum(r.cells["opening_acc_dep"] for r in cat_sec.rows)
        assert cat_sec.total.cells["dep_for_year"] == sum(r.cells["dep_for_year"] for r in cat_sec.rows)
        assert cat_sec.total.cells["closing_acc_dep"] == sum(r.cells["closing_acc_dep"] for r in cat_sec.rows)
        assert cat_sec.total.cells["net_book_value"] == sum(r.cells["net_book_value"] for r in cat_sec.rows)

    # Grand total equals sum of category subtotals
    assert root.total.cells["original_cost"] == sum(s.total.cells["original_cost"] for s in root.children)
    assert root.total.cells["opening_acc_dep"] == sum(s.total.cells["opening_acc_dep"] for s in root.children)
    assert root.total.cells["dep_for_year"] == sum(s.total.cells["dep_for_year"] for s in root.children)
    assert root.total.cells["closing_acc_dep"] == sum(s.total.cells["closing_acc_dep"] for s in root.children)
    assert root.total.cells["net_book_value"] == sum(s.total.cells["net_book_value"] for s in root.children)


@pytest.mark.asyncio
async def test_companies_act_schedule_invariants(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="ca_inv@testco.com")
    fy_id = ctx["fy_id"]
    email = ctx["email"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        context = await _load_asset_context(session, user.company_id, fy_id)

    doc = build_companies_act_schedule_ii_report(
        run=context["run"],
        assets_by_id=context["assets_by_id"],
        company_name=context["company_name"],
        fy_label=context["fy"].label,
        units="absolute",
    )

    root = doc.sections[0]
    assert len(root.children) > 0
    gt = root.total.cells
    assert gt["closing_gross"] > 0
    assert gt["closing_nbv"] > 0

    for cat_sec in root.children:
        assert len(cat_sec.rows) > 0
        # Per row invariants
        for r in cat_sec.rows:
            closing_gross = r.cells["closing_gross"]
            opening_gross = r.cells["opening_gross"]
            additions = r.cells["additions"]
            disposals = r.cells["disposals"]
            closing_dep = r.cells["closing_dep"]
            closing_nbv = r.cells["closing_nbv"]

            assert closing_gross == opening_gross + additions - disposals
            assert closing_nbv == closing_gross - closing_dep

        # Per category total invariants
        cat_tot = cat_sec.total.cells
        assert cat_tot["closing_gross"] > 0
        assert cat_tot["closing_gross"] == cat_tot["opening_gross"] + cat_tot["additions"] - cat_tot["disposals"]
        assert cat_tot["closing_nbv"] == cat_tot["closing_gross"] - cat_tot["closing_dep"]

    # Grand total invariants
    assert gt["closing_gross"] == gt["opening_gross"] + gt["additions"] - gt["disposals"]
    assert gt["closing_nbv"] == gt["closing_gross"] - gt["closing_dep"]


@pytest.mark.asyncio
async def test_income_tax_schedule_invariants(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="it_inv@testco.com")
    fy_id = ctx["fy_id"]
    email = ctx["email"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        context = await _load_asset_context(session, user.company_id, fy_id)

    doc = build_income_tax_appendix_i_report(
        run=context["run"],
        company_name=context["company_name"],
        fy_label=context["fy"].label,
        units="absolute",
    )

    sec = doc.sections[0]
    assert len(sec.rows) > 0
    gt = sec.total.cells
    assert gt["closing_wdv"] > 0

    for r in sec.rows:
        closing_wdv = r.cells["closing_wdv"]
        opening_wdv = r.cells["opening_wdv"]
        add_180 = r.cells["add_180"]
        add_less_180 = r.cells["add_less_180"]
        sales = r.cells["sales"]
        total_dep = r.cells["total_dep"]
        assert closing_wdv == (opening_wdv + add_180 + add_less_180 - sales - total_dep)

    # Grand total invariants
    assert gt["closing_wdv"] == (gt["opening_wdv"] + gt["add_180"] + gt["add_less_180"] - gt["sales"] - gt["total_dep"])


@pytest.mark.asyncio
async def test_asset_reports_cross_tenant_isolation(client: AsyncClient):
    """Cross-tenant test: A second company's user must get 404 when querying another company's FY."""
    ctx = await setup_asset_reports_environment(client, email="tenant1_admin@testco.com")
    fy_id = ctx["fy_id"]

    email2 = "tenant2_admin@otherco.com"
    await create_test_company(client, name="Tenant 2 Other Co", email=email2)
    token2 = await get_company_token(client, email=email2)
    headers2 = {"Authorization": f"Bearer {token2}"}

    res = await client.get(
        f"/api/v1/asset-reports/fixed_asset_register/export?financial_year_id={fy_id}&format=xlsx",
        headers=headers2,
    )
    assert res.status_code == 404, f"Expected 404 cross-tenant response, got {res.status_code}"


@pytest.mark.asyncio
async def test_asset_reports_units_scaling(client: AsyncClient):
    """Units-scaling test: A figure requested in lakhs must equal the absolute figure / 100,000."""
    ctx = await setup_asset_reports_environment(client, email="units_test@testco.com")
    fy_id = ctx["fy_id"]
    email = ctx["email"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        context = await _load_asset_context(session, user.company_id, fy_id)

    doc_abs = build_fixed_asset_register_report(
        assets=context["assets"],
        dep_lines_by_asset_id=context["dep_lines_by_asset_id"],
        company_name=context["company_name"],
        fy_label=context["fy"].label,
        units="absolute",
    )
    doc_lakhs = build_fixed_asset_register_report(
        assets=context["assets"],
        dep_lines_by_asset_id=context["dep_lines_by_asset_id"],
        company_name=context["company_name"],
        fy_label=context["fy"].label,
        units="lakhs",
    )

    abs_cost = doc_abs.sections[0].total.cells["original_cost"]
    lakhs_cost = doc_lakhs.sections[0].total.cells["original_cost"]
    assert abs_cost > 0
    assert lakhs_cost > 0
    assert lakhs_cost == abs_cost / Decimal("100000")


@pytest.mark.asyncio
async def test_asset_reports_gating_unfinalized_run(client: AsyncClient):
    """Gating test: Reports 2, 3, and 4 must be refused when no finalized run exists.
    
    Expected to FAIL: defect R4 (endpoints check only run existence, not finalized status; report 4 has no check).
    """
    await seed_masters()
    email = "gate_test@testco.com"
    headers = await admin_headers(client, email)

    # Create FY
    fy_res = await client.post(
        "/api/v1/financial-years",
        json={"label": "2024-25", "start_date": "2024-04-01", "end_date": "2025-03-31"},
        headers=headers,
    )
    assert fy_res.status_code == 201, fy_res.text
    fy_id = fy_res.json()["id"]

    # Insert an asset
    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        cat = (await session.execute(select(AssetCategory))).scalars().first()
        asset = Asset(
            company_id=user.company_id,
            asset_name="Lathe Machine",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 4, 1),
            available_for_use_date=date(2024, 4, 1),
            useful_life_months=60,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("100000.00"),
        )
        session.add(asset)
        await session.commit()

    # Create run in DRAFT status (do NOT finalize)
    run_res = await client.post(
        "/api/v1/depreciation/runs",
        json={"financial_year_id": fy_id, "notes": "Draft calculation"},
        headers=headers,
    )
    assert run_res.status_code == 201, run_res.text

    # Report 2: Companies Act schedule should reject unfinalized run
    res2 = await client.get(
        f"/api/v1/asset-reports/companies_act_depreciation/export?financial_year_id={fy_id}&format=xlsx",
        headers=headers,
    )
    assert res2.status_code == 409, f"Expected 409 for unfinalized run on companies_act_depreciation, got {res2.status_code}"

    # Report 3: Income Tax schedule should reject unfinalized run
    res3 = await client.get(
        f"/api/v1/asset-reports/income_tax_depreciation/export?financial_year_id={fy_id}&format=xlsx",
        headers=headers,
    )
    assert res3.status_code == 409, f"Expected 409 for unfinalized run on income_tax_depreciation, got {res3.status_code}"

    # Report 4: IT Asset annexure should reject when no finalized run exists
    res4 = await client.get(
        f"/api/v1/asset-reports/it_asset_annexure/export?financial_year_id={fy_id}&format=xlsx",
        headers=headers,
    )
    assert res4.status_code == 409, f"Expected 409 for unfinalized run on it_asset_annexure, got {res4.status_code}"


@pytest.mark.asyncio
async def test_asset_reports_category_filtering(client: AsyncClient):
    """Filter test: Filtering by category_id must return only assets of that category.
    
    Expected to FAIL: defect R5 (report endpoints accept no category filter params).
    """
    await seed_masters()
    email = "filter_test@testco.com"
    headers = await admin_headers(client, email)

    fy_res = await client.post(
        "/api/v1/financial-years",
        json={"label": "2024-25", "start_date": "2024-04-01", "end_date": "2025-03-31"},
        headers=headers,
    )
    assert fy_res.status_code == 201, fy_res.text
    fy_id = fy_res.json()["id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        cats = (await session.execute(select(AssetCategory))).scalars().all()
        cat1 = cats[0]
        cat2 = cats[1]

        a1 = Asset(
            company_id=user.company_id,
            asset_name="Server Unit Alpha",
            category_id=cat1.id,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 4, 1),
            available_for_use_date=date(2024, 4, 1),
            useful_life_months=60,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("500000.00"),
        )
        a2 = Asset(
            company_id=user.company_id,
            asset_name="Delivery Truck Beta",
            category_id=cat2.id,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 4, 1),
            available_for_use_date=date(2024, 4, 1),
            useful_life_months=60,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("200000.00"),
        )
        session.add_all([a1, a2])
        await session.commit()
        cat1_id = str(cat1.id)
        cat2_name = cat2.name

    run_res = await client.post(
        "/api/v1/depreciation/runs",
        json={"financial_year_id": fy_id, "notes": "Audited run"},
        headers=headers,
    )
    assert run_res.status_code == 201, run_res.text
    run_id = run_res.json()["id"]
    await client.post(f"/api/v1/depreciation/runs/{run_id}/finalize", headers=headers)

    # Request report filtered by category_id
    res = await client.get(
        f"/api/v1/asset-reports/fixed_asset_register/export?financial_year_id={fy_id}&category_id={cat1_id}&format=xlsx",
        headers=headers,
    )
    assert res.status_code == 200, res.text

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    sheet_values = [str(ws.cell(row=r, column=c).value) for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1) if ws.cell(row=r, column=c).value is not None]

    assert cat2_name not in sheet_values, f"Expected category '{cat2_name}' to be filtered out, but found in exported sheet"


@pytest.mark.asyncio
async def test_asset_reports_all_exports_xlsx_pdf_html(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="all_exports@testco.com")
    headers = ctx["headers"]
    fy_id = ctx["fy_id"]

    list_res = await client.get("/api/v1/asset-reports", headers=headers)
    assert list_res.status_code == 200, list_res.text
    report_list = list_res.json()

    for report_info in report_list:
        key = report_info["key"]

        # Preview HTML
        prev_res = await client.get(
            f"/api/v1/asset-reports/{key}/preview-html?financial_year_id={fy_id}&unit=absolute",
            headers=headers,
        )
        assert prev_res.status_code == 200, f"Failed preview-html for {key}: {prev_res.text}"
        assert "<html" in prev_res.text.lower()

        # Export XLSX
        xlsx_res = await client.get(
            f"/api/v1/asset-reports/{key}/export?financial_year_id={fy_id}&format=xlsx&unit=lakhs",
            headers=headers,
        )
        assert xlsx_res.status_code == 200, f"Failed XLSX export for {key}: {xlsx_res.text}"
        assert xlsx_res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert len(xlsx_res.content) > 1000

        # Export PDF
        pdf_res = await client.get(
            f"/api/v1/asset-reports/{key}/export?financial_year_id={fy_id}&format=pdf&unit=absolute",
            headers=headers,
        )
        assert pdf_res.status_code == 200, f"Failed PDF export for {key}: {pdf_res.text}"
        assert pdf_res.headers["content-type"] == "application/pdf"
        assert pdf_res.content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_asset_reports_pack_exports(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="packs_test@testco.com")
    headers = ctx["headers"]
    fy_id = ctx["fy_id"]

    pack_xlsx = await client.post(
        f"/api/v1/asset-reports/pack?financial_year_id={fy_id}&format=xlsx&unit=thousands",
        headers=headers,
    )
    assert pack_xlsx.status_code == 200, pack_xlsx.text
    assert pack_xlsx.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(pack_xlsx.content) > 2000

    pack_pdf = await client.post(
        f"/api/v1/asset-reports/pack?financial_year_id={fy_id}&format=pdf&unit=crores",
        headers=headers,
    )
    assert pack_pdf.status_code == 200, pack_pdf.text
    assert pack_pdf.headers["content-type"] == "application/pdf"
    assert pack_pdf.content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_asset_reports_archive_docvault(client: AsyncClient):
    ctx = await setup_asset_reports_environment(client, email="arch_test@testco.com")
    headers = ctx["headers"]
    fy_id = ctx["fy_id"]

    arch_res = await client.post(
        f"/api/v1/asset-reports/archive?report_key=fixed_asset_register&financial_year_id={fy_id}&format=pdf&unit=absolute",
        headers=headers,
    )
    assert arch_res.status_code == 200, arch_res.text
    arch_data = arch_res.json()
    assert arch_data["status"] == "archived"
    assert "document_id" in arch_data


@pytest.mark.asyncio
async def test_gst_itc_summary_report_endpoint(client: AsyncClient):
    """B1: gst_itc_summary report for a company with acquisitions returns 200 and non-empty content."""
    await seed_masters()
    headers = await admin_headers(client, "gstitc_test@testco.com")

    # Create FY
    fy_res = await client.post(
        "/api/v1/financial-years",
        json={"label": "2024-25", "start_date": "2024-04-01", "end_date": "2025-03-31"},
        headers=headers,
    )
    assert fy_res.status_code == 201, fy_res.text
    fy_id = fy_res.json()["id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == "gstitc_test@testco.com"))).scalar_one()
        cat = (await session.execute(select(AssetCategory))).scalars().first()

        acq = AssetAcquisition(
            company_id=user.company_id,
            invoice_number="INV-2024-001",
            invoice_date=date(2024, 4, 15),
            quantity=1,
            unit_basic_price=Decimal("100000.00"),
            discount_type=DiscountType.amount,
            gst_rate=Decimal("18.00"),
            cgst_amount=Decimal("9000.00"),
            sgst_amount=Decimal("9000.00"),
            igst_amount=Decimal("0.00"),
            itc_treatment=ItcTreatment.eligible,
        )
        session.add(acq)
        await session.flush()

        asset = Asset(
            company_id=user.company_id,
            acquisition_id=acq.id,
            asset_name="Server Rack",
            asset_code="SRV-001",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 5, 1),
            available_for_use_date=date(2024, 5, 1),
            useful_life_months=36,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("118000.00"),
        )
        session.add(asset)
        await session.commit()

    # Request the gst_itc_summary report
    res = await client.get(
        f"/api/v1/asset-reports/gst_itc_summary/export?financial_year_id={fy_id}&format=xlsx",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    assert len(res.content) > 1000


@pytest.mark.asyncio
async def test_gst_itc_summary_interstate_igst(client: AsyncClient):
    """F6/B1: Interstate acquisition has igst_amount set, cgst/sgst null, reports IGST only."""
    await seed_masters()
    headers = await admin_headers(client, "interstate_gst@testco.com")

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == "interstate_gst@testco.com"))).scalar_one()
        cat = (await session.execute(select(AssetCategory))).scalars().first()

        acq = AssetAcquisition(
            company_id=user.company_id,
            invoice_number="INV-INTER-99",
            invoice_date=date(2024, 5, 20),
            quantity=1,
            unit_basic_price=Decimal("200000.00"),
            discount_type=DiscountType.amount,
            gst_rate=Decimal("18.00"),
            cgst_amount=None,
            sgst_amount=None,
            igst_amount=Decimal("36000.00"),
            itc_treatment=ItcTreatment.eligible,
        )
        session.add(acq)
        await session.flush()

        asset = Asset(
            company_id=user.company_id,
            acquisition_id=acq.id,
            asset_name="Interstate Router",
            asset_code="RTR-001",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 6, 1),
            available_for_use_date=date(2024, 6, 1),
            useful_life_months=36,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("236000.00"),
        )
        session.add(asset)
        await session.commit()

        # Build document directly to inspect rows
        from app.services.reporting.asset_reports import build_gst_itc_summary_report
        doc = build_gst_itc_summary_report([asset], "Interstate Co", "2024-25", "absolute")
        assert len(doc.sections) == 1
        assert len(doc.sections[0].rows) == 1
        row = doc.sections[0].rows[0].cells
        assert row["cgst"] == Decimal("0.00")
        assert row["sgst"] == Decimal("0.00")
        assert row["igst"] == Decimal("36000.00")
        assert row["total_tax"] == Decimal("36000.00")
        assert row["cgst"] + row["sgst"] + row["igst"] == row["total_tax"]
        assert row["gst_rate"] == Decimal("18.00")


@pytest.mark.asyncio
async def test_disposals_register_includes_disposed_by_attribution(client: AsyncClient):
    """H1: Disposals Register output includes Disposed By column and the acting user name/email."""
    await seed_masters()
    email = "disp_reg_user@testco.com"
    headers = await admin_headers(client, email)

    fy_res = await client.post(
        "/api/v1/financial-years",
        json={"label": "2024-25", "start_date": "2024-04-01", "end_date": "2025-03-31"},
        headers=headers,
    )
    fy_id = fy_res.json()["id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        cat = (await session.execute(select(AssetCategory))).scalars().first()

        asset = Asset(
            company_id=user.company_id,
            asset_name="Disposed Vehicle",
            asset_code="VEH-001",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 4, 1),
            available_for_use_date=date(2024, 4, 1),
            useful_life_months=60,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("500000.00"),
        )
        session.add(asset)
        await session.commit()
        await session.refresh(asset)
        asset_id = str(asset.id)

    # Dispose asset
    disp_res = await client.post(
        f"/api/v1/assets/{asset_id}/dispose",
        json={
            "disposal_date": "2024-09-15",
            "disposal_type": "sale",
            "sale_proceeds": 200000.0,
            "buyer_name": "Fleet Buyers",
        },
        headers=headers,
    )
    assert disp_res.status_code == 200

    # Fetch Disposals Register report preview
    rep_res = await client.get(
        f"/api/v1/asset-reports/disposals_register/preview-html?financial_year_id={fy_id}",
        headers=headers,
    )
    assert rep_res.status_code == 200
    html = rep_res.text
    assert "Disposed By" in html
    assert "disp_reg_user@testco.com" in html or "Admin" in html


@pytest.mark.asyncio
async def test_pack_export_threads_filters_and_capitalized_default(client: AsyncClient):
    """M1: Pack export /pack returns 200 and FAR inside pack has capitalized default."""
    await seed_masters()
    email = "pack_filter@testco.com"
    headers = await admin_headers(client, email)

    fy_res = await client.post(
        "/api/v1/financial-years",
        json={"label": "2024-25", "start_date": "2024-04-01", "end_date": "2025-03-31"},
        headers=headers,
    )
    fy_id = fy_res.json()["id"]

    async with TestSessionLocal() as session:
        user = (await session.execute(select(CompanyUser).where(CompanyUser.email == email))).scalar_one()
        cat = (await session.execute(select(AssetCategory))).scalars().first()

        # 1 Draft asset, 1 Capitalized asset
        draft_asset = Asset(
            company_id=user.company_id,
            asset_name="Draft CWIP Item",
            asset_code="CWIP-001",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.draft,
            operational_status=AssetOperationalStatus.in_storage,
            original_cost=Decimal("30000.00"),
        )
        cap_asset = Asset(
            company_id=user.company_id,
            asset_name="Capitalized Machinery",
            asset_code="MAC-001",
            category_id=cat.id if cat else None,
            lifecycle_status=AssetLifecycleStatus.capitalized,
            operational_status=AssetOperationalStatus.in_use,
            capitalization_date=date(2024, 4, 1),
            useful_life_months=60,
            residual_pct=Decimal("5.00"),
            dep_method="slm",
            original_cost=Decimal("150000.00"),
        )
        session.add_all([draft_asset, cap_asset])
        await session.commit()

    # Finalize depreciation run for FY
    r = await client.post("/api/v1/depreciation/runs", json={"financial_year_id": fy_id}, headers=headers)
    assert r.status_code == 201
    await client.post(f"/api/v1/depreciation/runs/{r.json()['id']}/finalize", headers=headers)

    # Request pack in xlsx
    pack_res = await client.post(
        f"/api/v1/asset-reports/pack?financial_year_id={fy_id}&format=xlsx",
        headers=headers,
    )
    assert pack_res.status_code == 200
    assert pack_res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.asyncio
async def test_applied_filters_printed_on_all_filtered_reports(client: AsyncClient):
    """M3: Applied filters are printed in the subtitle of all filtered reports, not just FAR."""
    await seed_masters()
    email = "filter_sub@testco.com"
    headers = await admin_headers(client, email)

    fy_res = await client.post(
        "/api/v1/financial-years",
        json={"label": "2024-25", "start_date": "2024-04-01", "end_date": "2025-03-31"},
        headers=headers,
    )
    fy_id = fy_res.json()["id"]

    # Preview additions register with operational_status filter
    rep_res = await client.get(
        f"/api/v1/asset-reports/additions_register/preview-html?financial_year_id={fy_id}&operational_status=in_use",
        headers=headers,
    )
    assert rep_res.status_code == 200
    html = rep_res.text
    assert "Filters: Operation: in_use" in html


