"""Excel/CSV bulk import of pre-existing assets.

All-or-nothing: every row is validated before anything is written; one bad row
aborts the file with a per-row error report. Half-imported registers are
miserable to unwind; re-uploading a corrected sheet costs nothing.
"""
import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset_masters import AssetLookup, ItAssetBlock
from app.models.assets import Asset
from app.services.asset_existing import (
    ExistingAssetError,
    build_existing_asset,
    resolve_category_path,
)

IMPORT_HEADERS = [
    "Asset name", "Category", "Subcategory", "Original cost",
    "Purchase date", "Put-to-use date", "Capitalization date",
    "Opening accumulated depreciation", "Opening WDV (books)", "Opening WDV (tax)",
    "Useful life months", "Dep method", "Residual %", "IT block code",
    "Branch", "Location", "Department", "Cost centre", "Custodian name",
    "Serial number", "Remarks",
]

_INSTRUCTIONS = [
    "Fixed Asset Register — bulk import of pre-existing assets.",
    "Fill the Assets sheet. One row per asset. Do not rename columns.",
    "Dates: YYYY-MM-DD. Amounts: plain numbers, no currency symbols.",
    "Category/Subcategory must match your company's category names (case-insensitive).",
    "Assets dated before the current financial year MUST carry all three opening figures.",
    "Branch/Location/Department/Cost centre must match your master names.",
    "Any row with an error aborts the whole file — fix it and re-upload.",
]


class ImportRejected(Exception):
    def __init__(self, errors: List[dict]):
        self.errors = errors
        super().__init__("Import rejected")


class RowError(Exception):
    def __init__(self, row: int, message: str):
        self.row, self.message = row, message
        super().__init__(message)


def build_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    info = wb.active
    info.title = "Instructions"
    info.column_dimensions["A"].width = 95
    for i, line in enumerate(_INSTRUCTIONS, start=1):
        cell = info.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True)
    sheet = wb.create_sheet("Assets")
    for col, name in enumerate(IMPORT_HEADERS, start=1):
        sheet.cell(row=1, column=col, value=name).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(col)].width = 26
    sheet.append([
        "Tata Ace", "Motor vehicles", "Motor cars (other than those used in a hire business)",
        850000, "2022-06-10", "2022-06-20", "2022-06-30", 200000, 650000, 610000,
        None, None, None, None, None, None, None, None, "R Kumar", "DL1AB1234", "Example row — delete before use",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_date(value) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        raise ExistingAssetError(f"'{value}' is not a YYYY-MM-DD date")


def _to_dec(value) -> Optional[Decimal]:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        raise ExistingAssetError(f"'{value}' is not a valid amount")


def _to_int(value) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value)))
    except ValueError:
        raise ExistingAssetError(f"'{value}' is not a whole number")


def _cell(row: list, idx: int):
    return row[idx] if idx < len(row) else None


def parse_rows(rows: List[list]) -> List[dict]:
    """Header excluded by caller. Structural per-row parsing only; referential
    resolution against the DB happens in import_assets."""
    payloads: List[dict] = []
    for n, row in enumerate(rows, start=2):  # Excel numbering: header is row 1
        if not any(v not in (None, "") for v in row):
            continue  # tolerate blank spacer rows
        p: dict = {"row": n}

        def get(name: str, _row=row):
            return _cell(_row, IMPORT_HEADERS.index(name))

        try:
            name = get("Asset name")
            if not name or not str(name).strip():
                raise ExistingAssetError("Asset name is required")
            p["asset_name"] = str(name).strip()

            parent = get("Category")
            child = get("Subcategory")
            if not parent or not str(parent).strip():
                raise ExistingAssetError("Category is required")
            p["path"] = ([str(parent)] + ([str(child)] if child not in (None, "") else []))

            p["original_cost"] = _to_dec(get("Original cost"))
            if p["original_cost"] is None:
                raise ExistingAssetError("Original cost is required")

            p["purchase_date"] = _to_date(get("Purchase date"))
            p["put_to_use_date"] = _to_date(get("Put-to-use date"))
            p["capitalization_date"] = _to_date(get("Capitalization date"))
            p["opening_accumulated_depreciation"] = _to_dec(get("Opening accumulated depreciation"))
            p["opening_wdv"] = _to_dec(get("Opening WDV (books)"))
            p["opening_it_wdv"] = _to_dec(get("Opening WDV (tax)"))
            p["useful_life_months"] = _to_int(get("Useful life months"))
            p["residual_pct"] = _to_dec(get("Residual %"))
            for key in ("IT block code", "Branch", "Location", "Department",
                        "Cost centre", "Custodian name", "Serial number", "Remarks"):
                v = get(key)
                p[key.lower().replace(" ", "_")] = str(v).strip() if v not in (None, "") else None
        except ExistingAssetError as e:
            raise RowError(n, str(e))
        payloads.append(p)
    return payloads


async def _lookup_map(db: AsyncSession, company_id: uuid.UUID) -> dict:
    rows_ = (await db.execute(
        select(AssetLookup).where(AssetLookup.company_id == company_id)
    )).scalars().all()
    return {(str(l.kind.value), l.name.strip().lower()): l.id for l in rows_}


async def _it_block_map(db: AsyncSession, company_id: uuid.UUID) -> dict:
    rows_ = (await db.execute(
        select(ItAssetBlock).where(ItAssetBlock.company_id == company_id)
    )).scalars().all()
    return {b.code.strip().lower(): b.id for b in rows_}


async def import_assets(
    db: AsyncSession, company_id: uuid.UUID, user_id: uuid.UUID, rows: List[list]
) -> List[Asset]:
    payloads = parse_rows(rows)
    if not payloads:
        raise RowError(0, "No data rows found")

    lookups = await _lookup_map(db, company_id)
    blocks = await _it_block_map(db, company_id)

    errors: List[RowError] = []
    created: List[Asset] = []

    for p in payloads:
        try:
            branch_id = location_id = department_id = cost_centre_id = None
            for col, kind_key in (("branch", "branch"), ("location", "location"),
                                  ("department", "department"), ("cost_centre", "cost_centre")):
                raw = p.get(col)
                if raw:
                    lid = lookups.get((kind_key, raw.lower()))
                    if lid is None:
                        raise ExistingAssetError(f"Unknown {col.replace('_', ' ')} '{raw}'")
                    if col == "branch":
                        branch_id = lid
                    elif col == "location":
                        location_id = lid
                    elif col == "department":
                        department_id = lid
                    else:
                        cost_centre_id = lid

            block_code = p.get("it_block_code")
            it_block_id = None
            if block_code:
                it_block_id = blocks.get(block_code.lower())
                if it_block_id is None:
                    raise ExistingAssetError(f"Unknown IT block code '{block_code}'")

            category = await resolve_category_path(db, company_id, p["path"])
            unit = await build_existing_asset(
                db, company_id, user_id,
                asset_name=p["asset_name"],
                category=category,
                original_cost=p["original_cost"],
                purchase_date=p["purchase_date"],
                put_to_use_date=p["put_to_use_date"],
                capitalization_date=p["capitalization_date"],
                opening_accumulated_depreciation=p["opening_accumulated_depreciation"],
                opening_wdv=p["opening_wdv"],
                opening_it_wdv=p["opening_it_wdv"],
                useful_life_months=p["useful_life_months"],
                residual_pct=p["residual_pct"],
                branch_id=branch_id,
                location_id=location_id,
                department_id=department_id,
                cost_centre_id=cost_centre_id,
                custodian_name=p.get("custodian_name"),
                serial_number=p.get("serial_number"),
                remarks=p.get("remarks"),
                it_block_id=it_block_id,
            )
            created.append(unit)
        except ExistingAssetError as e:
            errors.append(RowError(p["row"], str(e)))

    if errors:
        await db.rollback()
        raise ImportRejected([{"row": e.row, "message": e.message} for e in errors])
    return created
