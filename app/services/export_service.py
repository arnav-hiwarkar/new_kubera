import io
from typing import List, Callable, Optional
from dataclasses import dataclass
import openpyxl

@dataclass
class ExportColumn:
    header: str
    field_path: str
    formatter: Optional[Callable] = None

def generate_xlsx(
    records: List[dict],
    columns: List[ExportColumn],
    sheet_name: str = "Export",
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    # Write headers
    for col_idx, col in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_idx, value=col.header)

    # Write data
    for row_idx, record in enumerate(records, start=2):
        for col_idx, col in enumerate(columns, start=1):
            
            # traverse field path (e.g. "custom_fields.serial_no")
            val = record
            for part in col.field_path.split('.'):
                if isinstance(val, dict):
                    val = val.get(part)
                else:
                    val = getattr(val, part, None)
                    
            if col.formatter and val is not None:
                try:
                    val = col.formatter(val)
                except:
                    pass
                    
            sheet.cell(row=row_idx, column=col_idx, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
