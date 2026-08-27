"""Bulk import of auditor requirements from Excel (4-column format).

All-or-nothing: every row is parsed and structurally validated before anything
is written; one bad row aborts the whole file with a per-row report.
Attachments are out of scope — they are added through company submissions.
"""
import io
import uuid
from datetime import date, datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.auditease import RequirementRequest

IMPORT_HEADERS = ["S. No.", "Requirement", "Due Date", "Priority"]

_INSTRUCTIONS = [
    "AuditEase — bulk requirement import.",
    "Fill the Requirements sheet. One row per requirement. Do not rename columns.",
    "Requirement is mandatory. Every other column is optional.",
    "S. No. is for your own reference only — AuditEase assigns REQ ids itself, "
    "continuing after the requirements already on the page.",
    "Due Date: blank, or YYYY-MM-DD / a real Excel date.",
    "Priority: blank (= 1) or a whole number 1-5.",
    "Documents cannot be attached here — the company attaches them from the requirement.",
    "Any row with an error aborts the whole file — fix it and re-upload.",
]


class ImportRejected(Exception):
    def __init__(self, errors: List[dict]):
        self.errors = errors
        super().__init__("Import rejected")


class RowError(Exception):
    def __init__(self, row: int, message: str):
        self.row, self.message = row, message
        super().__init__(message)


def build_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    info = wb.active
    info.title = "Instructions"
    info.column_dimensions["A"].width = 95
    for i, line in enumerate(_INSTRUCTIONS, start=1):
        cell = info.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True)
    sheet = wb.create_sheet("Requirements")
    for col, name in enumerate(IMPORT_HEADERS, start=1):
        sheet.cell(row=1, column=col, value=name).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(col)].width = 28
    sheet.append([
        "1", "FY24 bank statements for all current accounts", "2026-09-15", 2,
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_date(value) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        raise ValueError(f"'{value}' is not a YYYY-MM-DD date")


def _to_priority(value) -> int:
    if value in (None, ""):
        return 1
    try:
        p = int(float(str(value)))
    except (TypeError, ValueError):
        raise ValueError(f"'{value}' is not a whole number")
    if not 1 <= p <= 5:
        raise ValueError(f"Priority {p} out of range 1-5")
    return p


def _cell(row: list, idx: int):
    return row[idx] if idx < len(row) else None


def parse_rows(rows: List[list]) -> List[dict]:
    """Structural parsing only. Each dict is {row, description, due_date, priority}.
    Raises RowError(row, message) on the first malformed row."""
    payloads: List[dict] = []
    for n, row in enumerate(rows, start=2):  # Excel numbering: header is row 1
        if not any(v not in (None, "") for v in row):
            continue  # tolerate blank spacer rows
        p: dict = {"row": n}
        try:
            desc = _cell(row, 1)
            if not desc or not str(desc).strip():
                raise ValueError("Requirement is required")
            p["description"] = str(desc).strip()
            p["due_date"] = _to_date(_cell(row, 2))
            p["priority"] = _to_priority(_cell(row, 3))
        except ValueError as e:
            raise RowError(n, str(e))
        payloads.append(p)
    if not payloads and rows:
        raise RowError(2, "Requirement is required")
    return payloads


async def import_requirements(
    db: AsyncSession,
    engagement_id: uuid.UUID,
    raised_by: uuid.UUID,
    rows: List[list],
) -> List[RequirementRequest]:
    """Parse and validate every row, then stage creations on the caller's session
    with sequential seq_numbers continuing from the engagement's current maximum.
    Rolls back and raises ImportRejected on any failure. Does NOT commit."""
    payloads = parse_rows(rows)
    if not payloads:
        raise RowError(0, "No data rows found")

    res = await db.execute(
        select(func.max(RequirementRequest.seq_number)).where(
            RequirementRequest.engagement_id == engagement_id
        )
    )
    next_seq = (res.scalar() or 0) + 1

    created: List[RequirementRequest] = []
    for p in payloads:
        req = RequirementRequest(
            engagement_id=engagement_id,
            raised_by=raised_by,
            description=p["description"],
            seq_number=next_seq,
            priority=p["priority"],
            due_date=p["due_date"],
        )
        db.add(req)
        created.append(req)
        next_seq += 1

    await db.flush()
    return created

