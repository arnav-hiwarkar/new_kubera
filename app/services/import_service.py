import io
import csv
from decimal import Decimal
from uuid import UUID
from typing import List, Dict, Callable, Any
from dataclasses import dataclass, replace
from fastapi import UploadFile
import openpyxl

from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel, ValidationError

from app.models.custom_fields import CustomFieldDefinition
from app.services import trial_balance as tb
from app.services.custom_field_validator import validate_custom_fields

@dataclass
class ColumnMapping:
    source_column: str
    target_field: str

@dataclass
class ImportResult:
    imported: int
    skipped: int
    errors: List[dict]

async def parse_and_import(
    file: UploadFile,
    column_mappings: List[ColumnMapping],
    base_field_validators: Dict[str, Callable],
    custom_field_definitions: List[CustomFieldDefinition],
    row_factory: Callable[[dict, dict], Any], # takes base_data, custom_data
    db: AsyncSession,
    company_id: UUID,
    module: Any
) -> ImportResult:
    content = await file.read()
    
    rows = []
    if file.filename.endswith('.csv'):
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif file.filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(dict(zip(headers, row)))
    else:
        raise ValueError("Unsupported file format")

    result = ImportResult(imported=0, skipped=0, errors=[])
    
    def_map = {d.field_key: d for d in custom_field_definitions}

    for idx, row in enumerate(rows, start=1): # 1-indexed for user visibility
        row_errors = []
        base_data = {}
        custom_data = {}

        for mapping in column_mappings:
            source = mapping.source_column
            target = mapping.target_field
            value = row.get(source)

            if target in base_field_validators:
                try:
                    if value is not None and value != "":
                        base_data[target] = base_field_validators[target](value)
                except Exception as e:
                    row_errors.append(f"Invalid value for {target}: {value}")
            else:
                custom_data[target] = value

        # Validate custom fields
        custom_errors = await validate_custom_fields(custom_data, company_id, module, db)
        for err in custom_errors:
            row_errors.append(err)

        if row_errors:
            result.skipped += 1
            result.errors.append({"row": idx, "errors": row_errors})
        else:
            try:
                db_row = row_factory(base_data, custom_data)
                db.add(db_row)
                result.imported += 1
            except Exception as e:
                result.skipped += 1
                result.errors.append({"row": idx, "errors": [str(e)]})

    return result


# ---------------------------------------------------------------------------
# Trial-balance import (server-side: inspect -> map -> preview -> import)
#
# This module owns only file/sheet I/O and cell extraction. Every decision about
# what a number means -- its sign, its side, whether the row is consistent -- lives
# in app.services.trial_balance, which is pure and unit-tested there.
# ---------------------------------------------------------------------------

TB_TEXT_FIELDS = ("ledger_code", "ledger_name")
TB_NUMERIC_FIELDS = tb.NUMERIC_MAP_FIELDS


def _to_number(raw: Any) -> float:
    """Back-compat shim. Prefer tb.parse_amount, which keeps the Dr/Cr tag."""
    return float(tb.parse_amount(raw).value)


def _read_csv(content: bytes) -> List[list]:
    text = content.decode("utf-8-sig")
    return [list(r) for r in csv.reader(io.StringIO(text))]


def load_raw_rows(filename: str, content: bytes, sheet_name: str | None) -> List[list]:
    """Every row of one sheet, header included -- header detection happens later."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _read_csv(content)
    if name.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif sheet_name:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            else:
                ws = wb.worksheets[0]
            return [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    raise ValueError("Unsupported file format. Use .csv or .xlsx")


def load_sheet(
    filename: str, content: bytes, sheet_name: str | None,
    header_row: int | None = None, detect_header: bool = False,
) -> tuple[List[str], List[list]]:
    """Return (headers, data_rows) for one sheet. CSV has a single virtual sheet."""
    rows = load_raw_rows(filename, content, sheet_name)
    if not rows:
        return [], []
    hr = header_row if header_row is not None else (
        tb.detect_header_row(rows) if detect_header else 0
    )
    headers, first_data = tb.build_headers(rows, hr)
    return headers, rows[first_data:]


def inspect_spreadsheet(
    filename: str, content: bytes, preview: int = 5, detect_header: bool = False,
) -> List[dict]:
    """List every sheet with its headers and first `preview` data rows.

    `detect_header` is opt-in because the sales importer (the other caller) treats
    row 1 as the header unconditionally; turning detection on there would make its
    preview disagree with its import.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        raw_by_sheet = [("Sheet1", _read_csv(content))]
    elif name.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            raw_by_sheet = [(ws.title, [list(r) for r in ws.iter_rows(values_only=True)])
                            for ws in wb.worksheets]
        finally:
            wb.close()
    else:
        raise ValueError("Unsupported file format. Use .csv or .xlsx")

    sheets: List[dict] = []
    for title, rows in raw_by_sheet:
        hr = tb.detect_header_row(rows) if detect_header else 0
        headers, first_data = tb.build_headers(rows, hr)
        skipped = [
            [("" if c is None else str(c)) for c in row]
            for row in rows[:hr] if any(str(c or "").strip() for c in row)
        ]
        sheets.append({
            "name": title,
            "headers": headers,
            "preview_rows": [[("" if c is None else str(c)) for c in row]
                             for row in rows[first_data:first_data + preview]],
            "header_row": hr + 1,          # 1-indexed for humans
            "first_data_row": first_data + 1,
            "skipped_leading_rows": skipped,
            "suggested_map": tb.suggest_column_map(headers),
        })
    return sheets


# ---------------------------------------------------------------------------
# Parsing a trial balance
# ---------------------------------------------------------------------------

@dataclass
class ParsedTrialBalance:
    rows: List[dict]                      # DB-ready kwargs for TrialBalanceAccount
    errors: List[dict]                     # rows that could not be parsed
    dropped: List[dict]                    # blank / total / repeated-header rows
    validation: tb.TBValidation
    convention: tb.ConventionReport
    header_row: int                        # 0-indexed
    headers: List[str]
    sample_rows: List[dict]
    section_count: int = 0


def _cell(row: list, idx: Dict[str, int], field: str):
    i = idx.get(field)
    if i is None or i >= len(row):
        return None
    return row[i]


def parse_trial_balance(
    filename: str,
    content: bytes,
    sheet_name: str | None,
    column_map: Dict[str, Any],
    *,
    convention: tb.TBSignConvention | None = None,
    header_row: int | None = None,
    sample_size: int = 20,
) -> ParsedTrialBalance:
    """Parse the chosen sheet into canonical rows plus full diagnostics.

    Two passes: the first parses cells and infers the sign convention from the
    file as a whole (a single row cannot tell you whether credits are negative),
    the second normalizes with that convention. Row-level problems are collected,
    never raised -- only a structurally unusable mapping raises ValueError.
    """
    decimal_style = column_map.get("decimal_style") or "auto"
    credit_sign = column_map.get("credit_sign") or "auto"

    map_errors = tb.validate_column_map(column_map)
    if map_errors:
        raise ValueError("; ".join(map_errors))

    raw_rows = load_raw_rows(filename, content, sheet_name)
    if not raw_rows:
        raise ValueError("The sheet is empty")

    hr = header_row if header_row is not None else tb.detect_header_row(raw_rows)
    headers, first_data = tb.build_headers(raw_rows, hr)

    idx: Dict[str, int] = {}
    missing: List[str] = []
    for field, src in column_map.items():
        if not src or field in ("decimal_style", "credit_sign"):
            continue
        if src in headers:
            idx[field] = headers.index(src)
        else:
            missing.append(src)
    if missing:
        raise ValueError(f"Mapped columns not found in sheet: {missing}")
    if "ledger_name" not in idx:
        raise ValueError("ledger_name must be mapped")

    data_rows = raw_rows[first_data:]

    # --- pass 1: classify + parse cells ---
    errors: List[dict] = []
    dropped: List[dict] = []
    section_count = 0
    staged: List[tuple[int, str, str | None, Dict[str, tb.ParsedAmount]]] = []

    for offset, row in enumerate(data_rows):
        row_no = first_data + offset + 1  # 1-indexed source row, for the user
        kind = tb.classify_row(row, idx, headers)
        if kind in (tb.RowKind.blank, tb.RowKind.total, tb.RowKind.repeated_header):
            if kind is not tb.RowKind.blank:
                dropped.append({
                    "row": row_no,
                    "kind": kind.value,
                    "reason": _DROP_REASONS[kind],
                    "raw": [("" if c is None else str(c)) for c in row],
                })
            else:
                dropped.append({"row": row_no, "kind": kind.value,
                                "reason": _DROP_REASONS[kind], "raw": []})
            continue
        if kind is tb.RowKind.section:
            section_count += 1

        name = _cell(row, idx, "ledger_name")
        if name is None or str(name).strip() == "":
            errors.append({"row": row_no, "errors": ["ledger_name is required"]})
            continue

        code = None
        if "ledger_code" in idx:
            raw_code = _cell(row, idx, "ledger_code")
            code = None if raw_code in (None, "") else str(raw_code).strip()

        parsed: Dict[str, tb.ParsedAmount] = {}
        row_errors: List[str] = []
        for field in tb.NUMERIC_MAP_FIELDS:
            if field not in idx:
                continue
            raw = _cell(row, idx, field)
            try:
                parsed[field] = tb.parse_amount(raw, decimal_style=decimal_style)
            except tb.AmountParseError as e:
                row_errors.append(f"{field}: {e} (got {raw!r})")
        if row_errors:
            errors.append({"row": row_no, "errors": row_errors})
            continue
        staged.append((row_no, str(name).strip(), code, parsed))

    # --- infer the convention from the file as a whole ---
    has_closing_col = "closing_balance" in idx
    has_closing_pair = "closing_debit" in idx or "closing_credit" in idx
    closings = [p["closing_balance"] for _, _, _, p in staged if "closing_balance" in p]
    sum_dr = sum((p["debit"].magnitude for _, _, _, p in staged if "debit" in p), Decimal(0))
    sum_cr = sum((p["credit"].magnitude for _, _, _, p in staged if "credit" in p), Decimal(0))
    report = tb.detect_sign_convention(
        closings,
        has_closing_column=has_closing_col,
        has_closing_pair=has_closing_pair,
        sum_debit=sum_dr if "debit" in idx else None,
        sum_credit=sum_cr if "credit" in idx else None,
    )
    effective = convention
    override_evidence: str | None = None
    if effective is not None:
        override_evidence = "overridden by the user"
    elif report.convention in (tb.TBSignConvention.explicit, tb.TBSignConvention.derived):
        effective = report.convention
    elif credit_sign == "positive":
        effective = tb.TBSignConvention.magnitude
        override_evidence = "credit_sign=positive selected"
    elif credit_sign == "negative":
        effective = tb.TBSignConvention.signed
        override_evidence = "credit_sign=negative selected"
    else:
        effective = report.convention
    if override_evidence is not None:
        report = replace(report, convention=effective,
                         evidence=[*report.evidence, override_evidence])

    # --- pass 2: normalize ---
    out_rows: List[dict] = []
    figures: List[tb.RowFigures] = []
    for row_no, name, code, parsed in staged:
        amounts = tb.normalize_amounts(
            opening=parsed.get("opening_balance"),
            opening_debit=parsed.get("opening_debit"),
            opening_credit=parsed.get("opening_credit"),
            debit=parsed.get("debit"),
            credit=parsed.get("credit"),
            closing=parsed.get("closing_balance"),
            closing_debit=parsed.get("closing_debit"),
            closing_credit=parsed.get("closing_credit"),
            convention=effective,
            credit_sign=credit_sign,
            group_nature=None,   # mapping happens after import; see recanonicalize()
        )
        out_rows.append({
            "ledger_code": code,
            "ledger_name": name,
            "opening_balance": float(amounts.opening_balance),
            "debit": float(amounts.debit),
            "credit": float(amounts.credit),
            "closing_balance": float(amounts.closing_balance),
            "opening_net_debit": float(amounts.opening_net_debit),
            "closing_net_debit": float(amounts.closing_net_debit),
            "sign_unresolved": amounts.sign_unresolved,
            "source_row_consistent": amounts.row_consistent,
        })
        figures.append(tb.RowFigures(row=row_no, ledger_name=name, amounts=amounts))

    validation = tb.validate_rows(figures)
    sample = [
        {
            "row": f.row,
            "ledger_name": f.ledger_name,
            "opening_balance": float(f.amounts.opening_balance),
            "debit": float(f.amounts.debit),
            "credit": float(f.amounts.credit),
            "closing_balance": float(f.amounts.closing_balance),
            "closing_net_debit": float(f.amounts.closing_net_debit),
            "derived": list(f.amounts.derived),
            "notes": list(f.amounts.notes),
        }
        for f in figures[:sample_size]
    ]

    return ParsedTrialBalance(
        rows=out_rows,
        errors=errors,
        dropped=dropped,
        validation=validation,
        convention=report,
        header_row=hr,
        headers=headers,
        sample_rows=sample,
        section_count=section_count,
    )


_DROP_REASONS = {
    tb.RowKind.blank: "blank row",
    tb.RowKind.total: "total / carried-forward row",
    tb.RowKind.repeated_header: "repeated header row",
    tb.RowKind.section: "section heading",
}


def stated_totals(filename: str, content: bytes, sheet_name: str | None,
                  column_map: Dict[str, Any], header_row: int | None = None
                  ) -> tuple[float | None, float | None]:
    """The Dr/Cr figures off the sheet's own Total row, when it has one.

    A free cross-check: if the sheet says total debit is 5,000 and our own sum says
    something else, the user almost certainly mapped the wrong column.
    """
    raw_rows = load_raw_rows(filename, content, sheet_name)
    if not raw_rows:
        return None, None
    hr = header_row if header_row is not None else tb.detect_header_row(raw_rows)
    headers, first_data = tb.build_headers(raw_rows, hr)
    idx = {f: headers.index(s) for f, s in column_map.items()
           if s and f not in ("decimal_style", "credit_sign") and s in headers}
    for row in raw_rows[first_data:]:
        if tb.classify_row(row, idx, headers) is not tb.RowKind.total:
            continue
        try:
            dr = tb.parse_amount(_cell(row, idx, "debit")).magnitude
            cr = tb.parse_amount(_cell(row, idx, "credit")).magnitude
        except tb.AmountParseError:
            continue
        if dr or cr:
            return float(dr), float(cr)
    return None, None
