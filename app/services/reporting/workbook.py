"""Excel workbook renderer for neutral ReportDocument instances using openpyxl.

Produces styled, statutory-grade spreadsheets with:
- Standardized title block (company, report title, subtitle, period, units)
- Frozen header panes
- Distinct section headers and depth-first hierarchical sub-sections
- Thin top borders on sub-section totals (level >= 1) and double bottom borders on grand totals (level == 0)
- Auto-fitted column widths or explicit ColumnSpec widths
- Multi-sheet workbooks with an automated Cover sheet
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import Any, Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.reporting.document import (
    ColumnKind,
    ColumnSpec,
    ReportDocument,
    ReportRow,
    ReportSection,
    ReportTotal,
)
from app.services.reporting.format import format_date, format_percent, scale_for_units

# Font definitions
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="0F172A")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color="475569")
FONT_META = Font(name="Calibri", size=10, bold=True, color="334155")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="0F172A")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="1E293B")
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color="0F172A")
FONT_REGULAR = Font(name="Calibri", size=11, color="0F172A")
FONT_MUTED = Font(name="Calibri", size=11, color="64748B")
FONT_ITALIC = Font(name="Calibri", size=11, italic=True, color="334155")

# Fill definitions
FILL_HEADER = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
FILL_SECTION = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# Borders
BORDER_THIN = Side(style="thin", color="CBD5E1")
BORDER_SUBTOTAL = Border(
    top=Side(style="thin", color="0F172A"),
    bottom=Side(style="thin", color="E2E8F0"),
)
BORDER_GRAND_TOTAL = Border(
    top=Side(style="thin", color="0F172A"),
    bottom=Side(style="double", color="0F172A"),
)

# Number formats
NUM_FORMAT_MONEY = "#,##,##0.00"
NUM_FORMAT_NUMBER = "#,##,##0.00"
NUM_FORMAT_PERCENT = "0.00%"


def _format_cell_value(val: Any, kind: ColumnKind, units: str) -> tuple[Any, str | None]:
    """Return (python_val_or_formatted, openpyxl_number_format)."""
    if val is None or val == "":
        return "", None

    if kind == ColumnKind.money:
        if isinstance(val, (int, float, Decimal)):
            scaled = scale_for_units(val, units)
            return float(scaled), NUM_FORMAT_MONEY
        return str(val), None

    if kind == ColumnKind.number:
        if isinstance(val, (int, float, Decimal)):
            return float(val), NUM_FORMAT_NUMBER
        return str(val), None

    if kind == ColumnKind.percent:
        if isinstance(val, (int, float, Decimal)):
            return float(val) / 100.0 if abs(val) > 1 else float(val), NUM_FORMAT_PERCENT
        return str(val), None

    if kind == ColumnKind.date:
        return format_date(val), None

    return str(val), None


def _render_section(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    section: ReportSection,
    columns: tuple[ColumnSpec, ...],
    current_row: int,
    units: str,
    depth: int = 0,
) -> int:
    """Render a section and all its children depth-first."""
    num_cols = len(columns)

    # Section Title if present
    if section.title:
        cell = ws.cell(row=current_row, column=1, value=section.title)
        cell.font = FONT_SECTION
        cell.fill = FILL_SECTION
        ws.row_dimensions[current_row].height = 20
        if num_cols > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        current_row += 1

    # Data rows
    for r in section.rows:
        ws.row_dimensions[current_row].height = 18
        font = FONT_REGULAR
        if r.style == "muted":
            font = FONT_MUTED
        elif r.style in ("italic", "synthetic"):
            font = FONT_ITALIC

        for col_idx, col in enumerate(columns, start=1):
            val = r.cells.get(col.key)
            formatted_val, num_fmt = _format_cell_value(val, col.kind, units)
            cell = ws.cell(row=current_row, column=col_idx, value=formatted_val)
            cell.font = font
            if num_fmt:
                cell.number_format = num_fmt

            # Indentation on first column if applicable
            indent_level = r.indent + depth
            align = col.effective_align
            if col_idx == 1 and indent_level > 0:
                cell.alignment = Alignment(horizontal="left", indent=indent_level)
            else:
                cell.alignment = Alignment(horizontal=align)
        current_row += 1

    # Nested children
    for child in section.children:
        current_row = _render_section(
            ws, child, columns, current_row, units, depth=depth + 1
        )

    # Section Total
    if section.total:
        ws.row_dimensions[current_row].height = 19
        tot = section.total
        border = BORDER_GRAND_TOTAL if tot.level == 0 else BORDER_SUBTOTAL

        # Label in column 1
        lbl_cell = ws.cell(row=current_row, column=1, value=tot.label)
        lbl_cell.font = FONT_TOTAL
        lbl_cell.fill = FILL_TOTAL
        lbl_cell.border = border
        lbl_cell.alignment = Alignment(horizontal="left", indent=depth)

        # Other columns
        for col_idx, col in enumerate(columns[1:], start=2):
            val = tot.cells.get(col.key)
            formatted_val, num_fmt = _format_cell_value(val, col.kind, units)
            cell = ws.cell(row=current_row, column=col_idx, value=formatted_val)
            cell.font = FONT_TOTAL
            cell.fill = FILL_TOTAL
            cell.border = border
            if num_fmt:
                cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal=col.effective_align)

        current_row += 1

    return current_row


def render_document_to_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    doc: ReportDocument,
) -> None:
    """Render a single ReportDocument onto the given openpyxl worksheet."""
    # Find all columns across sections
    all_cols: list[ColumnSpec] = []
    if doc.sections:
        all_cols = list(doc.sections[0].columns)
    else:
        all_cols = [ColumnSpec(header="Particulars", key="particulars")]

    num_cols = len(all_cols)

    # 1. Title Block
    ws.cell(row=1, column=1, value=doc.company_name).font = FONT_TITLE
    ws.cell(row=2, column=1, value=doc.title).font = FONT_TITLE
    current_row = 3
    if doc.subtitle:
        ws.cell(row=current_row, column=1, value=doc.subtitle).font = FONT_SUBTITLE
        current_row += 1

    period_units_str = f"Period: {doc.period_label}   |   Amounts in: {doc.units.capitalize()}"
    ws.cell(row=current_row, column=1, value=period_units_str).font = FONT_META
    current_row += 2  # Blank line before table

    # 2. Table Headers
    header_row = current_row
    ws.row_dimensions[header_row].height = 24
    for col_idx, col in enumerate(all_cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col.header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal=col.effective_align, vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="0F172A"))

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    current_row += 1

    # 3. Sections
    for section in doc.sections:
        cols_to_use = section.columns or tuple(all_cols)
        current_row = _render_section(ws, section, cols_to_use, current_row, doc.units)

    # 4. Warnings / Footnotes at the bottom if present
    if doc.warnings:
        current_row += 1
        ws.cell(row=current_row, column=1, value="Notes / Warnings:").font = FONT_META
        current_row += 1
        for w in doc.warnings:
            ws.cell(row=current_row, column=1, value=f"• {w}").font = FONT_MUTED
            current_row += 1

    # 5. Column Widths
    for col_idx, col in enumerate(all_cols, start=1):
        col_letter = get_column_letter(col_idx)
        if col.width is not None:
            ws.column_dimensions[col_letter].width = max(col.width, 10)
        else:
            # Auto-fit width based on content length
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(header_row, current_row)
            ) if current_row > header_row else 12
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def render_cover_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheets: Sequence[tuple[str, ReportDocument]],
) -> None:
    """Render a clean Cover sheet for multi-sheet packs."""
    ws.title = "Cover"
    if not sheets:
        return

    first_doc = sheets[0][1]
    ws.cell(row=2, column=2, value=first_doc.company_name).font = Font(name="Calibri", size=18, bold=True)
    ws.cell(row=3, column=2, value="Financial & Statutory Reports Pack").font = Font(name="Calibri", size=14, bold=True, color="334155")
    ws.cell(row=5, column=2, value=f"Period: {first_doc.period_label}").font = FONT_META
    ws.cell(row=6, column=2, value=f"Reporting Currency Units: {first_doc.units.capitalize()}").font = FONT_META
    ws.cell(row=7, column=2, value=f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = FONT_MUTED

    ws.cell(row=9, column=2, value="Contents / Schedule of Reports:").font = FONT_HEADER
    r = 10
    for title, doc in sheets:
        ws.cell(row=r, column=2, value=f"{r - 9}. {title}").font = FONT_REGULAR
        ws.cell(row=r, column=3, value=doc.subtitle or doc.title).font = FONT_MUTED
        r += 1

    all_warnings = [w for _, d in sheets for w in d.warnings]
    if all_warnings:
        r += 1
        ws.cell(row=r, column=2, value="Audit Warnings / Exceptions:").font = FONT_META
        r += 1
        for w in all_warnings:
            ws.cell(row=r, column=2, value=f"• {w}").font = FONT_MUTED
            r += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 50


def write_document(doc: ReportDocument) -> io.BytesIO:
    """Write a single ReportDocument to an in-memory XLSX workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (doc.title[:28] if len(doc.title) > 28 else doc.title).replace("/", "-").replace(":", " ")
    render_document_to_sheet(ws, doc)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def write_workbook(sheets: list[tuple[str, ReportDocument]]) -> io.BytesIO:
    """Write multiple ReportDocument sheets with an automated Cover sheet into an in-memory XLSX."""
    wb = openpyxl.Workbook()
    cover_ws = wb.active
    render_cover_sheet(cover_ws, sheets)

    for sheet_name, doc in sheets:
        safe_name = (sheet_name[:28] if len(sheet_name) > 28 else sheet_name).replace("/", "-").replace(":", " ")
        ws = wb.create_sheet(title=safe_name)
        render_document_to_sheet(ws, doc)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
