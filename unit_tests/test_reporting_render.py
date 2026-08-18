"""Tests for rendering ReportDocument to XLSX and PDF formats."""
from decimal import Decimal
import io
import openpyxl
import pytest

from app.services.reporting.document import (
    ColumnKind,
    ColumnSpec,
    ReportDocument,
    ReportRow,
    ReportSection,
    ReportTotal,
)
from app.services.reporting.pdf import render_html, render_pdf, render_pack_pdf
from app.services.reporting.workbook import write_document, write_workbook


def _sample_document() -> ReportDocument:
    cols = (
        ColumnSpec(header="Particulars", key="particulars", kind=ColumnKind.text, width=30),
        ColumnSpec(header="Note", key="note_ref", kind=ColumnKind.text, width=10),
        ColumnSpec(header="Amount (INR)", key="amount", kind=ColumnKind.money, width=20),
    )

    sec1 = ReportSection(
        title="I. EQUITY AND LIABILITIES",
        columns=cols,
        rows=(
            ReportRow(cells={"particulars": "Share Capital", "note_ref": "1", "amount": Decimal("1000000.00")}),
            ReportRow(cells={"particulars": "Reserves & Surplus", "note_ref": "2", "amount": Decimal("500000.00")}),
        ),
        total=ReportTotal(label="Total Shareholders' Funds", cells={"amount": Decimal("1500000.00")}, level=1),
    )

    sec2 = ReportSection(
        title="II. ASSETS",
        columns=cols,
        rows=(
            ReportRow(cells={"particulars": "Property, Plant and Equipment", "note_ref": "3", "amount": Decimal("1200000.00")}),
            ReportRow(cells={"particulars": "Cash & Cash Equivalents", "note_ref": "4", "amount": Decimal("300000.00")}),
        ),
        total=ReportTotal(label="Total Assets", cells={"amount": Decimal("1500000.00")}, level=0),
    )

    return ReportDocument(
        title="Balance Sheet",
        subtitle="As at 31st March 2025 (Schedule III)",
        company_name="Test Enterprise Private Limited",
        period_label="2024-25",
        units="absolute",
        sections=(sec1, sec2),
        meta={"basis": "Indian GAAP (AS)"},
        warnings=("Engagement sign convention was verified.",),
    )


def test_render_xlsx():
    doc = _sample_document()
    stream = write_document(doc)
    assert isinstance(stream, io.BytesIO)
    stream.seek(0)

    wb = openpyxl.load_workbook(stream)
    ws = wb.active

    # Check title block
    assert ws.cell(row=1, column=1).value == "Test Enterprise Private Limited"
    assert ws.cell(row=2, column=1).value == "Balance Sheet"

    # Find the header row
    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "Particulars":
            header_row = r
            break
    assert header_row is not None
    assert ws.cell(row=header_row, column=2).value == "Note"
    assert ws.cell(row=header_row, column=3).value == "Amount (INR)"

    # Check that freeze panes is set
    assert ws.freeze_panes is not None

    # Check a data cell
    found_share_cap = False
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Share Capital":
            found_share_cap = True
            assert ws.cell(row=r, column=2).value == "1"
            assert ws.cell(row=r, column=3).value == 1000000.00
            break
    assert found_share_cap

    # Check total row has bold styling
    found_total_assets = False
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Total Assets":
            found_total_assets = True
            assert ws.cell(row=r, column=3).value == 1500000.00
            assert ws.cell(row=r, column=3).font.bold is True
            break
    assert found_total_assets


def test_render_multi_sheet_workbook():
    doc1 = _sample_document()
    doc2 = _sample_document()
    stream = write_workbook([("Balance Sheet", doc1), ("Profit and Loss", doc2)])
    assert isinstance(stream, io.BytesIO)
    stream.seek(0)

    wb = openpyxl.load_workbook(stream)
    assert "Cover" in wb.sheetnames
    assert "Balance Sheet" in wb.sheetnames
    assert "Profit and Loss" in wb.sheetnames


def test_render_pdf():
    doc = _sample_document()
    html = render_html(doc)
    assert "Test Enterprise Private Limited" in html
    assert "Balance Sheet" in html
    assert "10,00,000.00" in html or "1000000.00" in html

    pdf_bytes = render_pdf(doc)
    assert isinstance(pdf_bytes, bytes)
    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 1000


def test_render_pack_pdf():
    doc1 = _sample_document()
    doc2 = _sample_document()
    pdf_bytes = render_pack_pdf([doc1, doc2])
    assert isinstance(pdf_bytes, bytes)
    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 2000
