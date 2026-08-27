import io
from datetime import date, datetime
import openpyxl
import pytest

from app.services.requirement_import import (
    IMPORT_HEADERS, build_template_xlsx, parse_rows,
    RowError,
)


def _load(content: bytes) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    instructions = [row[0] for row in wb["Instructions"].iter_rows(values_only=True) if row and row[0]]
    ws = wb["Requirements"]
    req_rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    return instructions, req_rows


def test_template_has_headers_and_example():
    instructions, rows = _load(build_template_xlsx())
    assert len(instructions) >= 7
    assert list(rows[0]) == ["S. No.", "Requirement", "Due Date", "Priority"]
    assert len(rows[0]) == len(IMPORT_HEADERS)
    assert any(any(c for c in row) for row in rows[1:])  # example row present


def test_parse_minimal_row_defaults():
    # S. No. at index 0, Requirement at index 1
    p = parse_rows([["1", "Bank statements FY24", None, None]])[0]
    assert p["description"] == "Bank statements FY24"
    assert p["priority"] == 1
    assert p["due_date"] is None


def test_parse_full_row():
    p = parse_rows([["10", "Ledger dump", "2026-04-15", 3]])[0]
    assert p["description"] == "Ledger dump"
    assert p["due_date"] == date(2026, 4, 15)
    assert p["priority"] == 3


def test_parse_datetime_cell_accepted():
    p = parse_rows([["1", "Ledger dump", datetime(2026, 4, 15, 10, 30), "2"]])[0]
    assert p["due_date"] == date(2026, 4, 15)
    assert p["priority"] == 2


def test_missing_requirement_is_row_error():
    with pytest.raises(RowError) as e:
        parse_rows([["1", None, "2026-04-15", 1]])
    assert e.value.row == 2
    assert "Requirement is required" in e.value.message


def test_non_numeric_priority_rejected():
    with pytest.raises(RowError) as e:
        parse_rows([["1", "Valid requirement", None, "high"]])
    assert e.value.row == 2


def test_priority_out_of_range_rejected():
    with pytest.raises(RowError):
        parse_rows([["1", "Req", None, 0]])
    with pytest.raises(RowError):
        parse_rows([["1", "Req", None, 6]])


def test_malformed_date_rejected():
    with pytest.raises(RowError) as e:
        parse_rows([["1", "Req", "31/04/2025", 1]])
    assert e.value.row == 2


def test_blank_and_duplicate_s_no_ignored():
    payloads = parse_rows([
        [None, "Req One", None, 1],
        ["dup", "Req Two", None, 2],
        ["dup", "Req Three", None, 3],
    ])
    assert len(payloads) == 3
    assert payloads[0]["description"] == "Req One"
    assert payloads[1]["description"] == "Req Two"
    assert payloads[2]["description"] == "Req Three"


def test_blank_spacer_rows_tolerated():
    payloads = parse_rows([
        [None, None, None, None],
        ["1", "Real requirement", None, 1],
        [None, None, None, None],
    ])
    assert len(payloads) == 1
    assert payloads[0]["row"] == 3
    assert payloads[0]["description"] == "Real requirement"

